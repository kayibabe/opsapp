"""
rawdata_builder
===============

Step 1 of the Admin "Upload" tool — now **fiscal-year aware**.

The five zone workbooks (Liwonde, Mangochi, Mulanje, Ngabu, Zomba) for each
fiscal year live in their own sub-folder under ``dataupdater``:

    dataupdater/FY2324/<zone>.xlsx   → FY2023/24
    dataupdater/FY2425/<zone>.xlsx   → FY2024/25
    dataupdater/FY2526/<zone>.xlsx   → FY2025/26
    dataupdater/FY2627/<zone>.xlsx   → FY2026/27  (current)

``build_rawdata(year)`` compiles the chosen year's five workbooks into a clean
``RawData_updated.xlsx`` (DataEntry format) that Step 2 (preview/commit) imports.

This is a self-contained re-implementation of the legacy
``update_rawdata_master.py`` smart-diff script. Two correctness fixes baked in:

* It populates **Year** and **Fiscal Year** per fiscal month (the legacy script
  left them blank), so each year imports under the correct FY.
* It maps the source **TOTAL** rows (active customers, amount billed, cash
  collected, sales, …) — the legacy DataEntry master produced those via in-sheet
  formulas, so a freshly-built workbook would otherwise leave them 0 and badly
  understate billing/customer figures.
"""
from __future__ import annotations

import threading
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from app.core.config import BASE_DIR

DATAUPDATER_DIR = BASE_DIR / "dataupdater"
MASTER_TEMPLATE = DATAUPDATER_DIR / "RawData.xlsx"          # DataEntry header layout
OUTPUT_FILE = DATAUPDATER_DIR / "RawData_updated.xlsx"      # built workbook (Step 2 imports)

ZONES = ["Liwonde", "Mangochi", "Mulanje", "Ngabu", "Zomba"]

# FY end-year → (sub-folder name, FY label).  end-year 2027 = FY Apr-2026 … Mar-2027.
FY_FOLDERS: dict[int, tuple[str, str]] = {
    2024: ("FY2324", "FY2023/24"),
    2025: ("FY2425", "FY2024/25"),
    2026: ("FY2526", "FY2025/26"),
    2027: ("FY2627", "FY2026/27"),
}
CURRENT_FY = 2027

NON_DATA_SHEETS = {
    "Zone Monthly", "1st QTR", "2nd QTR ", "Midyear Summary", "3rd QTR", "4th QTR",
    "Zone Annual Summary", "Targets", "Budget", "ZA TH JA CALCULATION",
    "COST OF POWER", "WATER TARIFF",
}
MONTH_SOURCE_COLUMN = {"April": 3, "May": 4, "June": 5, "July": 6, "August": 7,
    "September": 8, "October": 9, "November": 10, "December": 11, "January": 12,
    "February": 13, "March": 14}
MONTH_NAME_TO_NO = {"April": 4, "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12, "January": 1,
    "February": 2, "March": 3}

# (source_row_index, dataentry_column_index) — component rows from the legacy map…
SRC_MAP: list[tuple[int, int]] = [
    (8, 7), (10, 8), (11, 9), (12, 10), (13, 11), (16, 13), (17, 14), (18, 15), (19, 16),
    (23, 19), (28, 21), (30, 23), (32, 25), (34, 27), (36, 29), (38, 31), (40, 33),
    (43, 35), (45, 37), (48, 39), (49, 40), (50, 41), (51, 42), (53, 43), (54, 44),
    (55, 45), (60, 49), (62, 51), (65, 52), (66, 53), (67, 54), (68, 55), (71, 58),
    (72, 59), (73, 60), (74, 61), (77, 64), (78, 65), (79, 66), (80, 67), (83, 70),
    (84, 71), (85, 72), (86, 73), (97, 80), (100, 81), (101, 82), (102, 83), (103, 84),
    (106, 86), (107, 87), (108, 88), (109, 89), (112, 91), (113, 92), (114, 93), (115, 94),
    (118, 97), (120, 98), (121, 99), (124, 101), (125, 102), (126, 103), (127, 104),
    (129, 106), (130, 107), (131, 108), (132, 109), (134, 111), (135, 112), (136, 113),
    (137, 114), (139, 116), (140, 117), (141, 118), (142, 119), (152, 126), (153, 127),
    (154, 128), (155, 129), (156, 130), (157, 131), (158, 132), (159, 133), (160, 134),
    (161, 135), (162, 136), (163, 137), (164, 138), (166, 139), (167, 140), (168, 141),
    (169, 142), (170, 143), (171, 144), (172, 145), (173, 146), (174, 147), (176, 148),
    (177, 149), (178, 150), (179, 151), (180, 152), (181, 153), (183, 154), (184, 155),
    (185, 156), (186, 157), (188, 158), (189, 159), (190, 160), (191, 161), (193, 163),
    (194, 164), (195, 165), (196, 166), (198, 167), (199, 168), (200, 169), (201, 170),
    (202, 171), (205, 173), (206, 174), (207, 175), (208, 176), (211, 178), (212, 179),
    (213, 180), (214, 181), (218, 184), (219, 185), (220, 186), (221, 187), (224, 189),
    (225, 190), (226, 191), (227, 192), (231, 195), (232, 196), (233, 197), (234, 198),
    (237, 200), (238, 201), (239, 202), (240, 203), (244, 206), (245, 207), (252, 212),
    (253, 213), (254, 214), (255, 215), (256, 216), (260, 219), (261, 220),
    # …plus the TOTAL rows the legacy DataEntry computed by formula:
    (192, 162),  # Total Pipe Breakdowns (→ pipe_breakdowns)
    (14, 12),    # TOTAL Vol Billed Postpaid
    (20, 17),    # TOTAL Vol Billed Prepaid
    (21, 18),    # TOTAL Revenue Water  (→ revenue_water)
    (57, 46),    # Total Operating Costs (→ op_cost)
    (246, 208),  # Total Debtors (→ total_debtors)
    (110, 90),   # TOTAL Active Postpaid
    (116, 95),   # TOTAL Active Prepaid
    (117, 96),   # TOTAL Active Customers
    (209, 177),  # TOTAL Cash Coll PP
    (215, 182),  # TOTAL Cash Coll Prepaid
    (216, 183),  # TOTAL Cash Collected
    (222, 188),  # TOTAL Amt Billed PP
    (228, 193),  # TOTAL Amt Billed Prepaid
    (229, 194),  # TOTAL Amount Billed
    (242, 205),  # TOTAL Sales MWK
    # ── Previously missing fields (audit fix) ───────────────────────────
    (122, 100),  # % Population Supplied    → pct_pop_supplied
    (258, 218),  # Connectivity Rate        → connectivity_rate
    (257, 217),  # Connection Days          → connection_days
    (262, 221),  # Response Time avg        → response_time_avg
    (89,  76),   # ALL Conn BroughtFwd      → all_conn_bfwd
    (94,  79),   # ALL Conn CarriedFwd      → all_conn_cfwd
    (70,  57),   # Indiv Conn CarriedFwd    → conn_indiv_cfwd
    (76,  63),   # Inst Conn CarriedFwd     → conn_inst_cfwd
    (82,  69),   # Comm Conn CarriedFwd     → conn_comm_cfwd
    (88,  75),   # CWP Conn CarriedFwd      → conn_cwp_cfwd
    (128, 105),  # Inst StuckM CarriedFwd   → stuck_inst_cfwd
    (133, 110),  # Comm StuckM CarriedFwd   → stuck_comm_cfwd
    (138, 115),  # CWP StuckM CarriedFwd    → stuck_cwp_cfwd
    (143, 120),  # Indiv StuckM CarriedFwd  → stuck_indiv_cfwd
    (148, 125),  # ALL StuckM CarriedFwd    → all_stuck_cfwd
]

_BUILD_LOCK = threading.Lock()


class BuildError(RuntimeError):
    """Raised when source files or the master template cannot be located."""


def _quarter(mn: int) -> str:
    return "Q1" if mn in (4, 5, 6) else "Q2" if mn in (7, 8, 9) else "Q3" if mn in (10, 11, 12) else "Q4"


def _num(v) -> float:
    try:
        if pd.isna(v):
            return 0.0
    except Exception:
        pass
    if isinstance(v, str):
        v = v.strip().replace(",", "")
        if v == "":
            return 0.0
    try:
        return float(v)
    except Exception:
        return 0.0


def _derive_folder(year: int) -> str:
    """Derive FY folder name from end-year: 2027 → 'FY2627', 2028 → 'FY2728'."""
    return f"FY{(year - 1) % 100:02d}{year % 100:02d}"


def _derive_label(year: int) -> str:
    return f"FY{year - 1}/{str(year)[-2:]}"


def resolve_year(year: int | None) -> int:
    y = int(year) if year else CURRENT_FY
    if y < 2020 or y > 2040:
        raise BuildError(f"Fiscal year end-year {y} is outside the supported range (2020–2040)")
    return y


def fy_dir(year: int) -> Path:
    y = resolve_year(year)
    folder = FY_FOLDERS[y][0] if y in FY_FOLDERS else _derive_folder(y)
    return DATAUPDATER_DIR / folder


def output_file_path() -> Path:
    return OUTPUT_FILE


def available_years() -> list[dict]:
    """List the fiscal years the tool can build, with per-year readiness, for the UI."""
    out = []
    for end_year, (folder, label) in sorted(FY_FOLDERS.items(), reverse=True):
        statuses = zone_files_status(end_year)
        present = sum(1 for s in statuses if s["exists"])
        out.append({
            "year": end_year,
            "label": label,
            "folder": folder,
            "files_present": present,
            "files_total": len(ZONES),
            "ready": present == len(ZONES),
            "current": end_year == CURRENT_FY,
        })
    return out


def zone_files_status(year: int | None = None) -> list[dict]:
    """Report which zone workbooks are present in the given FY's sub-folder."""
    d = fy_dir(year)
    statuses = []
    for zone in ZONES:
        p = d / f"{zone}.xlsx"
        statuses.append({
            "zone": zone,
            "file": f"{FY_FOLDERS[resolve_year(year)][0]}/{zone}.xlsx",
            "exists": p.exists(),
            "modified": p.stat().st_mtime if p.exists() else None,
        })
    return statuses


def build_rawdata(year: int | None = None, *, test_mode: bool = False,
                  force_overwrite: bool = False) -> dict:
    """Compile the selected fiscal year's five zone workbooks into RawData_updated.xlsx."""
    end_year = resolve_year(year)
    if end_year in FY_FOLDERS:
        folder, label = FY_FOLDERS[end_year]
    else:
        folder = _derive_folder(end_year)
        label  = _derive_label(end_year)
    src_dir = fy_dir(end_year)

    if not MASTER_TEMPLATE.exists():
        raise BuildError(f"Master template not found: {MASTER_TEMPLATE}")
    missing = [z for z in ZONES if not (src_dir / f"{z}.xlsx").exists()]
    if missing:
        raise FileNotFoundError(
            f"Missing zone file(s) for {label}: {', '.join(missing)} — place them in dataupdater/{folder}."
        )

    with _BUILD_LOCK:
        mwb = load_workbook(MASTER_TEMPLATE, read_only=True)
        ws = mwb["DataEntry"]
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        headers = list(row2)
        ncol = len(headers)
        mwb.close()

        rows: list[dict] = []
        files_read = 0
        skipped_empty = 0
        errors = 0
        log_lines: list[str] = [f"Building {label} from dataupdater/{folder}"]

        for zone in ZONES:
            zpath = src_dir / f"{zone}.xlsx"
            try:
                zx = pd.ExcelFile(zpath)
            except Exception as exc:
                errors += 1
                log_lines.append(f"ERROR opening {zone}: {exc}")
                continue
            files_read += 1
            for sheet in zx.sheet_names:
                if sheet in NON_DATA_SHEETS:
                    continue
                try:
                    sdf = pd.read_excel(zpath, sheet_name=sheet, header=None)
                except Exception as exc:
                    errors += 1
                    log_lines.append(f"ERROR reading {zone}/{sheet}: {exc}")
                    continue
                for mname, scol in MONTH_SOURCE_COLUMN.items():
                    if scol >= sdf.shape[1]:
                        continue
                    probe = sdf.iloc[8, scol] if sdf.shape[0] > 8 else None
                    blank = (probe is None
                             or (isinstance(probe, str) and probe.strip() == "")
                             or (not isinstance(probe, str) and _num(probe) == 0))
                    if blank:
                        skipped_empty += 1
                        continue
                    mn = MONTH_NAME_TO_NO[mname]
                    rec = {i: None for i in range(ncol)}
                    rec[0] = zone
                    rec[1] = sheet
                    rec[2] = label
                    rec[3] = end_year - 1 if mn >= 4 else end_year
                    rec[4] = mn
                    rec[5] = mname
                    rec[6] = _quarter(mn)
                    for srow, rcol in SRC_MAP:
                        if rcol >= ncol:
                            continue
                        try:
                            rec[rcol] = _num(sdf.iloc[srow, scol])
                        except Exception:
                            rec[rcol] = 0.0
                    rows.append(rec)

        summary = {
            "source_files_read": files_read,
            "rows_added": len(rows),
            "rows_updated": 0,
            "skipped_already_complete": 0,
            "skipped_empty_source": skipped_empty,
            "errors": errors,
            "schemes": len({(r[0], r[1]) for r in rows}),
            "filled_records": [],
        }

        if not test_mode:
            out = Workbook()
            o = out.active
            o.title = "DataEntry"
            for j, v in enumerate(row1, start=1):
                o.cell(row=1, column=j, value=v)
            for j, v in enumerate(headers, start=1):
                o.cell(row=2, column=j, value=v)
            for ri, rec in enumerate(rows, start=3):
                for ci in range(ncol):
                    o.cell(row=ri, column=ci + 1, value=rec[ci])
            out.save(OUTPUT_FILE)
            log_lines.append(f"Wrote {len(rows)} rows to {OUTPUT_FILE.name}")

    return {
        "ok": errors == 0,
        "exit_code": 0 if errors == 0 else 1,
        "year": end_year,
        "fy_label": label,
        "test_mode": test_mode,
        "force_overwrite": force_overwrite,
        "records_changed": len(rows),
        "output_file": OUTPUT_FILE.name,
        "output_exists": OUTPUT_FILE.exists() and not test_mode,
        "summary": summary,
        "log": log_lines,
    }
