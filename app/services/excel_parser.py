"""
services/excel_parser.py
========================
Parses SRWB RawData Excel uploads. Produces a structured ParseResult
containing validated rows, type-coerced metrics, statistical anomaly
flags, and conflict markers against the live database.

The only public API you need:
    parser = ExcelParser()
    result = parser.parse(file_obj, db_conn)   # pure read — no DB writes

Design notes
------------
- All DB interaction is read-only (anomaly history + conflict lookup).
- No Flask context is used — this module is framework-agnostic.
- The parser never raises on bad data; every issue is captured as a
  RowIssue with severity "error" or "warning" on the ParsedRow.
  Only truly unrecoverable format problems (wrong file type, missing
  required dimension columns) raise ValueError to the caller.
"""

from __future__ import annotations

import re
import sqlite3
import logging
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl

log = logging.getLogger(__name__)

# ── Column map: normalised Excel header  →  DB column name ───────────────
#
# "Normalised" means: lowercased, punctuation stripped, spaces → underscores.
# Add synonyms here whenever the template evolves. The parser never cares
# about capitalisation or minor punctuation differences in the Excel header.
#
COLUMN_MAP: dict[str, str] = {
    # ── DIMENSIONS (row 2 of DataEntry sheet, normalised) ────────────────
    "zone":                              "zone",
    "scheme":                            "scheme",
    "fiscal_year":                       "fiscal_year",
    "year":                              "year",
    "month_no":                          "month_no",
    "month":                             "month",
    "quarter":                           "quarter",

    # ── WATER PRODUCTION & NRW ───────────────────────────────────────────
    # "Volume Produced (m³)"  →  "volume_produced_m"
    "volume_produced_m":                 "vol_produced",
    "vol_produced":                      "vol_produced",          # legacy alias
    # "Vol Billed Individual Postpaid"  →  "vol_billed_individual_postpaid"
    "vol_billed_individual_postpaid":    "vol_billed_indiv_pp",
    "vol_billed_cwp_postpaid":           "vol_billed_cwp_pp",
    "vol_billed_institutions_postpaid":  "vol_billed_inst_pp",
    "vol_billed_commercial_postpaid":    "vol_billed_comm_pp",
    "total_vol_billed_postpaid":         "total_vol_billed_pp",
    "vol_billed_individual_prepaid":     "vol_billed_indiv_prepaid",
    "vol_billed_cwp_prepaid":            "vol_billed_cwp_prepaid",
    "vol_billed_institutions_prepaid":   "vol_billed_inst_prepaid",
    "vol_billed_commercial_prepaid":     "vol_billed_comm_prepaid",
    "total_vol_billed_prepaid":          "total_vol_billed_prepaid",
    # "TOTAL Revenue Water m³"  →  "total_revenue_water_m"
    "total_revenue_water_m":             "revenue_water",
    # "Non-Revenue Water m³"   →  "non_revenue_water_m"
    "non_revenue_water_m":               "nrw",
    # "% NRW"  →  "nrw"  (% stripped, normalise leaves "nrw")
    "nrw":                               "pct_nrw",

    # ── TREATMENT CHEMICALS ───────────────────────────────────────────────
    "chlorine_kg":                       "chlorine_kg",
    # "Alum Sulphate kg"  →  "alum_sulphate_kg"
    "alum_sulphate_kg":                  "alum_kg",
    "soda_ash_kg":                       "soda_ash_kg",
    "algae_floc_litres":                 "algae_floc_litres",
    "sud_floc_litres":                   "sud_floc_litres",
    # "Potassium Permanganate kg"  →  "potassium_permanganate_kg"
    "potassium_permanganate_kg":         "kmno4_kg",
    # "Cost of Chemicals MWK"  →  "cost_of_chemicals_mwk"
    "cost_of_chemicals_mwk":             "chem_cost",
    # "Chem Cost per m³"  →  "chem_cost_per_m"
    "chem_cost_per_m":                   "chem_cost_per_m3",

    # ── POWER ─────────────────────────────────────────────────────────────
    # "Power Usage kWh"  →  "power_usage_kwh"
    "power_usage_kwh":                   "power_kwh",
    # "Cost of Power MWK"  →  "cost_of_power_mwk"
    "cost_of_power_mwk":                 "power_cost",
    # "Power Cost per m³"  →  "power_cost_per_m"
    "power_cost_per_m":                  "power_cost_per_m3",

    # ── TRANSPORT & OPERATIONS ────────────────────────────────────────────
    # "Distances Covered km"  →  "distances_covered_km"
    "distances_covered_km":              "distances_km",
    "fuel_used_litres":                  "fuel_used_litres",
    # "Cost of Fuel MWK"  →  "cost_of_fuel_mwk"
    "cost_of_fuel_mwk":                  "fuel_cost",
    # "Maintenance MWK"  →  "maintenance_mwk"
    "maintenance_mwk":                   "maintenance",
    # "Staff Costs MWK"  →  "staff_costs_mwk"
    "staff_costs_mwk":                   "staff_costs",
    # "Wages MWK"  →  "wages_mwk"
    "wages_mwk":                         "wages",
    # "Other Overhead MWK"  →  "other_overhead_mwk"
    "other_overhead_mwk":                "other_overhead",
    # "TOTAL Operating Costs MWK"  →  "total_operating_costs_mwk"
    "total_operating_costs_mwk":         "op_cost",
    # "OpCost per m³ Produced"  →  "opcost_per_m_produced"
    "opcost_per_m_produced":             "op_cost_per_m3_produced",
    # "OpCost per m³ Billed"  →  "opcost_per_m_billed"
    "opcost_per_m_billed":               "op_cost_per_m3_billed",

    # ── STAFFING ──────────────────────────────────────────────────────────
    "permanent_staff":                   "perm_staff",
    "temporary_staff":                   "temp_staff",

    # ── CONNECTIONS — AGGREGATED (individual types not stored in DB) ──────
    # "ALL Conn BroughtFwd"  →  "all_conn_broughtfwd"
    "all_conn_broughtfwd":               "all_conn_bfwd",
    "all_conn_applied":                  "all_conn_applied",
    # "ALL Conn TOTAL Done"  →  "all_conn_total_done"
    "all_conn_total_done":               "new_connections",
    # "ALL Conn CarriedFwd"  →  "all_conn_carriedfwd"
    "all_conn_carriedfwd":               "all_conn_cfwd",
    "prepaid_meters_installed":          "prepaid_meters_installed",

    # ── DISCONNECTIONS ────────────────────────────────────────────────────
    "disconnected_individual":           "disconnected_individual",
    # "Disconnected Institutional"  →  "disconnected_institutional"
    "disconnected_institutional":        "disconnected_inst",
    "disconnected_commercial":           "disconnected_commercial",
    "disconnected_cwp":                  "disconnected_cwp",
    "total_disconnected":                "total_disconnected",

    # ── ACTIVE CONSUMERS — POSTPAID ───────────────────────────────────────
    # "Active Postpaid Individual"  →  "active_postpaid_individual"
    "active_postpaid_individual":        "active_post_individual",
    "active_postpaid_institutional":     "active_post_inst",
    "active_postpaid_commercial":        "active_post_commercial",
    "active_postpaid_cwp":               "active_post_cwp",
    "total_active_postpaid":             "active_postpaid",

    # ── ACTIVE CONSUMERS — PREPAID ────────────────────────────────────────
    "active_prepaid_individual":         "active_prep_individual",
    "active_prepaid_institutional":      "active_prep_inst",
    "active_prepaid_commercial":         "active_prep_commercial",
    "active_prepaid_cwp":                "active_prep_cwp",
    "total_active_prepaid":              "active_prepaid",

    # ── ACTIVE TOTALS ─────────────────────────────────────────────────────
    "total_active_customers":            "active_customers",
    "total_metered_consumers":           "total_metered",

    # ── POPULATION ────────────────────────────────────────────────────────
    "population_supply_area":            "pop_supply_area",
    "population_supplied":               "pop_supplied",
    "pct_population_supplied":           "pct_pop_supplied",

    # ── STUCK METERS — AGGREGATED (by-type not stored in DB) ─────────────
    # "ALL StuckM BroughtFwd"  →  "all_stuckm_broughtfwd"
    "all_stuckm_broughtfwd":             "stuck_meters",
    "all_stuckm_new":                    "stuck_new",
    "all_stuckm_repaired":               "stuck_repaired",
    "all_stuckm_replaced":               "stuck_replaced",
    # CarriedFwd not in DB — intentionally skipped

    # ── PIPE BREAKDOWNS (individual sizes summed into material totals) ────
    # "TOTAL Pipe Breakdowns"  →  "total_pipe_breakdowns"
    "total_pipe_breakdowns":             "pipe_breakdowns",
    # PVC sizes are stored individually and also rolled into pipe_pvc
    "pvc_20mm":                         "pvc_20mm",
    "pvc_25mm":                         "pvc_25mm",
    "pvc_32mm":                         "pvc_32mm",
    "pvc_40mm":                         "pvc_40mm",
    "pvc_50mm":                         "pvc_50mm",
    "pvc_63mm":                         "pvc_63mm",
    "pvc_75mm":                         "pvc_75mm",
    "pvc_90mm":                         "pvc_90mm",
    "pvc_110mm":                        "pvc_110mm",
    "pvc_160mm":                        "pvc_160mm",
    "pvc_200mm":                        "pvc_200mm",
    "pvc_250mm":                        "pvc_250mm",
    "pvc_315mm":                        "pvc_315mm",
    # Non-PVC breakdown sizes are aggregated to their stored material totals.
    "gi_15mm":                          "_gi_15mm",
    "gi_20mm":                          "_gi_20mm",
    "gi_25mm":                          "_gi_25mm",
    "gi_40mm":                          "_gi_40mm",
    "gi_50mm":                          "_gi_50mm",
    "gi_75mm":                          "_gi_75mm",
    "gi_100mm":                         "_gi_100mm",
    "gi_150mm":                         "_gi_150mm",
    "gi_200mm":                         "_gi_200mm",
    "di_150mm":                         "_di_150mm",
    "di_200mm":                         "_di_200mm",
    "di_250mm":                         "_di_250mm",
    "di_300mm":                         "_di_300mm",
    "di_350mm":                         "_di_350mm",
    "di_525mm":                         "_di_525mm",
    "hdpe_20mm":                        "_hdpe_20mm",
    "hdpe_25mm":                        "_hdpe_25mm",
    "hdpe_32mm":                        "_hdpe_32mm",
    "hdpe_50mm":                        "_hdpe_50mm",
    "ac_50mm":                          "_ac_50mm",
    "ac_75mm":                          "_ac_75mm",
    "ac_100mm":                         "_ac_100mm",
    "ac_150mm":                         "_ac_150mm",

    # ── PUMPS & SUPPLY HOURS ──────────────────────────────────────────────
    "pump_breakdowns":                   "pump_breakdowns",
    "pump_hours_lost":                   "pump_hours_lost",
    # "Normal Supply Hours"  →  "normal_supply_hours"
    "normal_supply_hours":               "supply_hours",
    # "Power Failure Hours"  →  "power_failure_hours"
    "power_failure_hours":               "power_fail_hours",

    # ── DEVELOPMENT LINES ─────────────────────────────────────────────────
    "devlines_32mm":                     "dev_lines_32mm",
    "devlines_50mm":                     "dev_lines_50mm",
    "devlines_63mm":                     "dev_lines_63mm",
    "devlines_90mm":                     "dev_lines_90mm",
    "devlines_110mm":                    "dev_lines_110mm",
    # "TOTAL Dev Lines Done"  →  "total_dev_lines_done"
    "total_dev_lines_done":              "dev_lines_total",

    # ── CASH COLLECTED (aggregates only) ──────────────────────────────────
    "total_cash_coll_pp":                "cash_coll_pp",
    "total_cash_coll_prepaid":           "cash_coll_prepaid",
    "total_cash_collected":              "cash_collected",

    # ── AMOUNTS BILLED (aggregates only) ─────────────────────────────────
    "total_amt_billed_pp":               "amt_billed_pp",
    "total_amt_billed_prepaid":          "amt_billed_prepaid",
    "total_amount_billed":               "amt_billed",

    # ── SERVICE CHARGES & METER RENTAL ───────────────────────────────────
    "total_service_charge":              "service_charge",
    "total_meter_rental":                "meter_rental",
    # "TOTAL Sales MWK"  →  "total_sales_mwk"
    "total_sales_mwk":                   "total_sales",

    # ── DEBTORS ───────────────────────────────────────────────────────────
    # "Private Debtors MWK"  →  "private_debtors_mwk"
    "private_debtors_mwk":               "private_debtors",
    "public_debtors_mwk":                "public_debtors",
    "total_debtors_mwk":                 "total_debtors",

    # ── FINANCIAL KPIs ────────────────────────────────────────────────────
    # "OpCost per Sales"  →  "opcost_per_sales"
    "opcost_per_sales":                  "op_cost_per_sales",
    # "Cash Collection Rate"  →  "cash_collection_rate"
    "cash_collection_rate":              "collection_rate",
    # "Collection per Total Sales"  →  "collection_per_total_sales"
    "collection_per_total_sales":        "collection_per_sales",

    # ── CONNECTION PERFORMANCE ────────────────────────────────────────────
    # "Cust Applied Connection"  →  "cust_applied_connection"
    "cust_applied_connection":           "conn_applied",
    "days_to_quotation":                 "days_to_quotation",
    # "Cust Fully Paid"  →  "cust_fully_paid"
    "cust_fully_paid":                   "conn_fully_paid",
    "days_to_connect":                   "days_to_connect",
    "connectivity_rate":                 "connectivity_rate",

    # ── QUERY PERFORMANCE ─────────────────────────────────────────────────
    "queries_received":                  "queries_received",
    # "Time to Resolve Queries"  →  "time_to_resolve_queries"
    "time_to_resolve_queries":           "time_to_resolve",
    "response_time_avg":                 "response_time_avg",
}

# Columns that must be present and non-null on every row
REQUIRED_DIMS: tuple[str, ...] = ("zone", "scheme", "month", "year")

# Metric columns where a null value is a hard error (row excluded)
REQUIRED_METRICS: tuple[str, ...] = ("vol_produced",)

# Metrics used in anomaly detection.
# CRITICAL: these must match the actual column names in the `records` DB
# table (Record ORM in database.py), NOT the parser's internal aliases.
# The anomaly query runs against the live DB, so wrong names → SQL error.
ANOMALY_METRICS: tuple[str, ...] = (
    "vol_produced",       # DB column: vol_produced     ✓
    "active_customers",   # DB column: active_customers  (was: cust_active ✗)
    "pct_nrw",            # DB column: pct_nrw           (was: nrw_pct ✗)
    "amt_billed",         # DB column: amt_billed         (was: total_billed ✗)
    "cash_collected",     # DB column: cash_collected     (was: total_collections ✗)
    "new_connections",    # DB column: new_connections    (was: nwc_done ✗)
)

# A value is anomalous if it's outside this band relative to 3-month average
ANOMALY_HIGH = 3.0    # 3× average → too high
ANOMALY_LOW  = 0.33   # ⅓ of average → too low

MONTH_NAMES: dict[str, int] = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}

# Reverse map: integer month number → full month name as stored in the DB.
# The DB stores month as a string (e.g. "April"), while ParsedRow.month is
# an integer (1–12).  Use this when building queries against the live DB.
MONTH_INT_TO_NAME: dict[int, str] = {
    1: "January", 2: "February", 3: "March",    4: "April",
    5: "May",     6: "June",     7: "July",      8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


# ── Data classes ──────────────────────────────────────────────────────────

@dataclass
class RowIssue:
    severity: str   # "error" | "warning"
    field: str
    message: str

    def to_dict(self) -> dict:
        return {"severity": self.severity, "field": self.field, "message": self.message}


@dataclass
class ParsedRow:
    row_num:  int
    zone:     str
    scheme:   str
    month:    int
    year:     int
    metrics:  dict[str, Any]
    status:   str = "ok"    # "ok" | "warning" | "error"
    issues:   list[RowIssue] = field(default_factory=list)
    conflict: Optional[dict] = None   # populated by _detect_conflicts

    @property
    def key(self) -> tuple[str, str, int, int]:
        return (self.zone, self.scheme, self.month, self.year)

    def add_issue(self, severity: str, field_name: str, message: str) -> None:
        self.issues.append(RowIssue(severity, field_name, message))
        if severity == "error":
            self.status = "error"
        elif severity == "warning" and self.status == "ok":
            self.status = "warning"

    def to_dict(self) -> dict:
        return {
            "row_num":  self.row_num,
            "zone":     self.zone,
            "scheme":   self.scheme,
            "month":    self.month,
            "year":     self.year,
            "metrics":  self.metrics,
            "status":   self.status,
            "issues":   [i.to_dict() for i in self.issues],
            "conflict": self.conflict,
        }


@dataclass
class ParseResult:
    rows:                    list[ParsedRow] = field(default_factory=list)
    unrecognised_columns:    list[str]       = field(default_factory=list)
    missing_required_columns: list[str]      = field(default_factory=list)
    period_month:            Optional[int]   = None
    period_year:             Optional[int]   = None

    @property
    def importable_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.status != "error"]

    @property
    def error_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.status == "error"]

    @property
    def conflict_rows(self) -> list[ParsedRow]:
        return [r for r in self.importable_rows if r.conflict is not None]

    def to_dict(self) -> dict:
        return {
            "total_rows":              len(self.rows),
            "importable_count":        len(self.importable_rows),
            "error_count":             len(self.error_rows),
            "warning_count":           sum(1 for r in self.rows if r.status == "warning"),
            "conflict_count":          len(self.conflict_rows),
            "period_month":            self.period_month,
            "period_year":             self.period_year,
            "unrecognised_columns":    self.unrecognised_columns,
            "missing_required_columns": self.missing_required_columns,
            "rows":                    [r.to_dict() for r in self.rows],
        }


# ── Core parser class ─────────────────────────────────────────────────────

class ExcelParser:
    """
    One instance per upload. Call parse() once; it returns a ParseResult.
    All DB access is SELECT-only.
    """

    def parse(self, file_obj, db_conn: sqlite3.Connection) -> ParseResult:
        """
        Full pipeline:
          open workbook → find sheet → read & map headers
          → parse rows → coerce types → detect anomalies → detect conflicts
        
        Raises ValueError for fatal format problems (bad file, missing
        required dimension columns). All row-level problems are captured
        as RowIssues — they never raise.
        """
        wb = self._open_workbook(file_obj)
        ws = self._find_data_sheet(wb)
        col_map, unrecognised, missing_required = self._read_headers(ws)

        result = ParseResult(
            unrecognised_columns=unrecognised,
            missing_required_columns=missing_required,
        )

        if missing_required:
            raise ValueError(
                f"Required columns not found in sheet: "
                f"{', '.join(missing_required)}. "
                f"Check the file matches the SRWB template and re-upload."
            )

        for row_num, raw_row in enumerate(
            ws.iter_rows(min_row=3, values_only=True), start=3
        ):
            if all(cell is None for cell in raw_row):
                continue
            parsed = self._parse_row(raw_row, row_num, col_map)
            result.rows.append(parsed)

        self._infer_period(result)
        self._detect_anomalies(result, db_conn)
        self._detect_conflicts(result, db_conn)

        log.info(
            "Parse complete: %d rows total, %d importable, %d errors, %d conflicts",
            len(result.rows),
            len(result.importable_rows),
            len(result.error_rows),
            len(result.conflict_rows),
        )
        return result

    # ── Workbook helpers ──────────────────────────────────────────────────

    def _open_workbook(self, file_obj) -> openpyxl.Workbook:
        try:
            return openpyxl.load_workbook(
                file_obj, read_only=True, data_only=True
            )
        except Exception as exc:
            raise ValueError(f"Cannot open workbook: {exc}") from exc

    def _find_data_sheet(self, wb: openpyxl.Workbook):
        """
        Preferred sheet names in priority order.
        Falls back to the first (active) sheet.
        """
        preferred = ("DataEntry", "dataentry", "rawdata", "data", "sheet1", "records")
        lower_map = {name.lower(): name for name in wb.sheetnames}
        for p in preferred:
            if p in lower_map:
                return wb[lower_map[p]]
        return wb.active

    # ── Header processing ─────────────────────────────────────────────────

    @staticmethod
    def _normalize_header(raw: Any) -> str:
        """
        'Vol. Produced (m³)'  →  'vol_produced_m'
        'NRW %'               →  'nrw_'          (trailing _ is harmless)
        'NWCs B/F'            →  'nwcs_bf'

        Steps (order matters):
          1. Encode to ASCII, dropping non-ASCII glyphs (strips ³, ², etc.)
          2. Remove slashes/backslashes without leaving a space ('B/F' → 'BF')
          3. Replace remaining punctuation with a space
          4. Collapse whitespace to underscores
        """
        if raw is None:
            return ""
        s = str(raw).strip().lower()
        s = s.encode("ascii", "ignore").decode("ascii")  # drop ³, ², etc.
        s = re.sub(r"[/\\]", "", s)                      # /\ → nothing
        s = re.sub(r"[^\w\s]", " ", s)                   # punctuation → space
        s = re.sub(r"\s+", "_", s.strip())
        s = re.sub(r"_+", "_", s)
        return s.strip("_")

    def _read_headers(
        self, ws
    ) -> tuple[dict[int, str], list[str], list[str]]:
        """
        Returns:
          col_map           – {col_index: db_column_name}
          unrecognised      – raw header strings not in COLUMN_MAP
          missing_required  – required DB dim columns absent from the sheet
        """
        # DataEntry sheet has TWO header rows:
        #   Row 1: section group labels (WATER PRODUCTION & NRW, STAFFING …)
        #   Row 2: actual column names (Zone, Scheme, Volume Produced (m³) …)
        # We must read row 2 for the column name map.
        header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))

        col_map: dict[int, str] = {}
        unrecognised: list[str] = []

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            normalised = self._normalize_header(cell)
            if normalised in COLUMN_MAP:
                db_col = COLUMN_MAP[normalised]
                # Last mapping wins if duplicate headers exist
                col_map[idx] = db_col
            else:
                unrecognised.append(str(cell).strip())

        mapped_db_cols = set(col_map.values())
        missing_required = [d for d in REQUIRED_DIMS if d not in mapped_db_cols]

        return col_map, unrecognised, missing_required

    # ── Row parsing ───────────────────────────────────────────────────────

    def _parse_row(
        self,
        raw_row: tuple,
        row_num: int,
        col_map: dict[int, str],
    ) -> ParsedRow:
        # Extract every mapped column from the raw tuple
        raw: dict[str, Any] = {
            db_col: (raw_row[idx] if idx < len(raw_row) else None)
            for idx, db_col in col_map.items()
        }

        # Parse dimensions first
        zone   = self._coerce_str(raw.get("zone"))
        scheme = self._coerce_str(raw.get("scheme"))
        month  = self._coerce_month(raw.get("month"))
        year   = self._coerce_year(raw.get("year"))

        # Separate metrics from dimensions
        metrics: dict[str, Any] = {
            k: v for k, v in raw.items() if k not in REQUIRED_DIMS
        }

        parsed = ParsedRow(
            row_num=row_num,
            zone=zone or "",
            scheme=scheme or "",
            month=month or 0,
            year=year or 0,
            metrics=metrics,
        )

        # ── Validate dimensions ──────────────────────────────────────────
        if not zone:
            parsed.add_issue("error", "zone", "Zone is missing or blank.")
        if not scheme:
            parsed.add_issue("error", "scheme", "Scheme is missing or blank.")
        if month is None:
            parsed.add_issue(
                "error", "month",
                f"Cannot parse month from value: '{raw.get('month')}'."
            )
        if year is None:
            parsed.add_issue(
                "error", "year",
                f"Cannot parse year from value: '{raw.get('year')}'."
            )

        # ── Validate required metrics ────────────────────────────────────
        for req in REQUIRED_METRICS:
            val = metrics.get(req)
            if val is None:
                parsed.add_issue(
                    "error", req,
                    f"'{req}' is required but missing for "
                    f"{zone or '?'} / {scheme or '?'}."
                )

        # ── Coerce all metric values to float (or None) ──────────────────
        for col, val in list(metrics.items()):
            if val is None:
                continue
            coerced = self._coerce_numeric(val)
            if coerced is None and str(val).strip() not in ("", "-", "—"):
                parsed.add_issue(
                    "warning", col,
                    f"Non-numeric value '{val}' for column '{col}' — "
                    f"treated as missing."
                )
            metrics[col] = coerced

        self._rollup_breakdown_metrics(metrics)

        return parsed

    @staticmethod
    def _rollup_breakdown_metrics(metrics: dict[str, Any]) -> None:
        """
        Roll temporary per-size pipe-breakdown fields into the stored DB columns.
        - PVC sizes remain stored individually and also contribute to pipe_pvc.
        - GI, DI, HDPE and AC sizes are aggregated into their material totals.
        - Temporary fields are removed so commit/import only sees valid DB columns.
        """
        def _sum(fields: list[str]) -> float:
            total = 0.0
            for field in fields:
                val = metrics.get(field)
                if isinstance(val, (int, float)):
                    total += float(val)
            return total

        pvc_fields = [
            'pvc_20mm','pvc_25mm','pvc_32mm','pvc_40mm','pvc_50mm','pvc_63mm',
            'pvc_75mm','pvc_90mm','pvc_110mm','pvc_160mm','pvc_200mm','pvc_250mm','pvc_315mm',
        ]
        gi_fields = ['_gi_15mm','_gi_20mm','_gi_25mm','_gi_40mm','_gi_50mm','_gi_75mm','_gi_100mm','_gi_150mm','_gi_200mm']
        di_fields = ['_di_150mm','_di_200mm','_di_250mm','_di_300mm','_di_350mm','_di_525mm']
        hdpe_ac_fields = ['_hdpe_20mm','_hdpe_25mm','_hdpe_32mm','_hdpe_50mm','_ac_50mm','_ac_75mm','_ac_100mm','_ac_150mm']

        pvc_total = _sum(pvc_fields)
        gi_total = _sum(gi_fields)
        di_total = _sum(di_fields)
        hdpe_ac_total = _sum(hdpe_ac_fields)

        if pvc_total or 'pipe_pvc' not in metrics or metrics.get('pipe_pvc') is None:
            metrics['pipe_pvc'] = pvc_total
        if gi_total or 'pipe_gi' not in metrics or metrics.get('pipe_gi') is None:
            metrics['pipe_gi'] = gi_total
        if di_total or 'pipe_di' not in metrics or metrics.get('pipe_di') is None:
            metrics['pipe_di'] = di_total
        if hdpe_ac_total or 'pipe_hdpe_ac' not in metrics or metrics.get('pipe_hdpe_ac') is None:
            metrics['pipe_hdpe_ac'] = hdpe_ac_total

        for field in gi_fields + di_fields + hdpe_ac_fields:
            metrics.pop(field, None)

    # ── Type coercion helpers ─────────────────────────────────────────────

    @staticmethod
    def _coerce_str(val: Any) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    @staticmethod
    def _coerce_numeric(val: Any) -> Optional[float]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        cleaned = re.sub(r"[,\s]", "", str(val).strip())
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _coerce_month(val: Any) -> Optional[int]:
        """Accept integer 1–12, month names, or Excel date objects."""
        if val is None:
            return None
        if hasattr(val, "month"):           # datetime from openpyxl
            return val.month
        if isinstance(val, (int, float)):
            m = int(val)
            return m if 1 <= m <= 12 else None
        s = str(val).strip().lower()
        if s.isdigit():
            m = int(s)
            return m if 1 <= m <= 12 else None
        return MONTH_NAMES.get(s)

    @staticmethod
    def _coerce_year(val: Any) -> Optional[int]:
        """Accept 2024 or 24 (→ 2024), or Excel date objects."""
        if val is None:
            return None
        if hasattr(val, "year"):
            return val.year
        try:
            y = int(float(str(val).strip()))
            return 2000 + y if y < 100 else y
        except (ValueError, TypeError):
            return None

    # ── Period inference ──────────────────────────────────────────────────

    @staticmethod
    def _infer_period(result: ParseResult) -> None:
        """Detect the dominant (month, year) across all parsed rows."""
        months = [r.month for r in result.rows if r.month and r.month > 0]
        years  = [r.year  for r in result.rows if r.year  and r.year  > 0]
        if months:
            result.period_month = max(set(months), key=months.count)
        if years:
            result.period_year = max(set(years), key=years.count)

    # ── Anomaly detection ─────────────────────────────────────────────────

    def _detect_anomalies(
        self, result: ParseResult, db_conn: sqlite3.Connection
    ) -> None:
        """
        For each ANOMALY_METRIC, compare the incoming value against the
        trailing 3-month average for that (zone, scheme) in the DB.
        Flag if outside the [ANOMALY_LOW, ANOMALY_HIGH] band.

        Rows with fewer than 2 historical records are skipped — not enough
        history to establish a meaningful baseline.
        """
        cur = db_conn.cursor()
        metric_cols = ", ".join(ANOMALY_METRICS)

        for row in result.importable_rows:
            # Use month_no (INTEGER column) for ordering and comparison —
            # NOT the `month` column which is a string ("April", "January"…).
            # Comparing an integer row.month against a string DB column
            # produces silently wrong results in SQLite.
            cur.execute(
                f"""
                SELECT {metric_cols}
                FROM   records
                WHERE  zone     = ?
                  AND  scheme   = ?
                  AND  (year < ? OR (year = ? AND month_no < ?))
                ORDER  BY year DESC, month_no DESC
                LIMIT  3
                """,
                (row.zone, row.scheme, row.year, row.year, row.month),
            )
            history = cur.fetchall()
            if len(history) < 2:
                continue

            for m_idx, metric in enumerate(ANOMALY_METRICS):
                new_val = row.metrics.get(metric)
                if not isinstance(new_val, (int, float)):
                    continue

                hist_vals = [
                    h[m_idx] for h in history
                    if h[m_idx] is not None
                ]
                if not hist_vals:
                    continue

                avg = sum(hist_vals) / len(hist_vals)
                if avg == 0:
                    continue

                ratio = new_val / avg

                if ratio > ANOMALY_HIGH:
                    row.add_issue(
                        "warning", metric,
                        f"Value {new_val:,.1f} is {ratio:.1f}× the "
                        f"{len(hist_vals)}-month trailing average "
                        f"({avg:,.1f}). Verify before importing."
                    )
                elif ratio < ANOMALY_LOW:
                    row.add_issue(
                        "warning", metric,
                        f"Value {new_val:,.1f} is only {ratio:.0%} of the "
                        f"{len(hist_vals)}-month trailing average "
                        f"({avg:,.1f}). Possible data-entry error."
                    )

    # ── Conflict detection ────────────────────────────────────────────────

    def _detect_conflicts(
        self, result: ParseResult, db_conn: sqlite3.Connection
    ) -> None:
        """
        Batch-fetch all (zone, scheme, month, year) combos that already
        exist in the DB for the importable rows. Attach a conflict dict to
        each row that has an existing record so the caller can surface the
        old vs new values for the operator to resolve.

        Two bugs fixed here vs the original:
          1. Column names now match the actual DB schema (active_customers,
             pct_nrw, amt_billed, cash_collected) not the parser's internal
             aliases (cust_active, nrw_pct, total_billed, total_collections).
          2. ParsedRow.month is an integer (1-12); the DB stores month as a
             string ("April"…"March").  We convert using MONTH_INT_TO_NAME
             before building the query and key the result dict accordingly.
        """
        importable = result.importable_rows
        if not importable:
            return

        cur = db_conn.cursor()

        # Convert each row's integer month to the string the DB stores
        # e.g. row.month = 4  →  month_str = "April"
        def _month_str(row: ParsedRow) -> str:
            return MONTH_INT_TO_NAME.get(row.month, str(row.month))

        # Build a single batch query rather than N individual queries
        placeholders = ", ".join("(?, ?, ?, ?)" for _ in importable)
        params: list[Any] = []
        for row in importable:
            params.extend([row.zone, row.scheme, _month_str(row), row.year])

        cur.execute(
            f"""
            SELECT zone, scheme, month, year,
                   active_customers, vol_produced, pct_nrw,
                   amt_billed, cash_collected
            FROM   records
            WHERE  (zone, scheme, month, year) IN ({placeholders})
            """,
            params,
        )

        # Key by (zone, scheme, month_string, year) for O(1) lookup
        existing: dict[tuple, dict] = {
            (r[0], r[1], r[2], r[3]): {
                "active_customers": r[4],
                "vol_produced":     r[5],
                "pct_nrw":          r[6],
                "amt_billed":       r[7],
                "cash_collected":   r[8],
            }
            for r in cur.fetchall()
        }

        for row in importable:
            # Look up using the string month — matches what we inserted above
            lookup_key = (row.zone, row.scheme, _month_str(row), row.year)
            ex = existing.get(lookup_key)
            if ex:
                row.conflict = {
                    "existing": ex,
                    "incoming": {
                        k: row.metrics.get(k)
                        for k in (
                            "active_customers", "vol_produced", "pct_nrw",
                            "amt_billed", "cash_collected",
                        )
                    },
                }
