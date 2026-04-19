"""
main.py — FastAPI application entry point.

Run with:
    uvicorn app.main:app --reload --port 8000

Authentication
--------------
All /api/* routes (except /api/auth/login) require a JWT in the header:
    Authorization: Bearer <token>

Role-based access is enforced via FastAPI dependencies at the router level:
  - All read endpoints        → get_current_user  (any valid role)
  - /api/upload/*             → require_admin     (admin only)
  - /api/records/export/csv   → require_export    (admin or user; not viewer)
  - /api/admin/*              → require_admin     (admin only)
"""
from contextlib import asynccontextmanager

from fastapi import Depends, FastAPI, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
import os
import time
import uuid

from app.auth import ensure_default_admin, get_current_user, require_admin
from app.core.config import settings
from app.core.logging import REQUEST_ID_CTX, logger as app_logger
from app.database import SessionLocal, create_tables
from app.routers import analytics, budget, catalogue, panels, records, reports, upload, insights
from app.routers.users import _limiter, admin_router, auth_router
from slowapi import _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: create DB tables, bootstrap default admin if needed."""
    settings.validate_startup()
    create_tables()
    db = SessionLocal()
    try:
        ensure_default_admin(db)
        # Auto-import: if records table is empty and an Excel file exists, seed it
        from app.database import Record
        count = db.query(Record).count()
        if count == 0:
            _auto_import(db)
    finally:
        db.close()
    print("[OK] Database tables ready")
    print("[OK] User authentication active  (JWT / bcrypt)")
    yield


def _auto_import(db):
    """Seed DB from RawData.xlsx if available. Called only when records table is empty."""
    import glob
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = (
        glob.glob(os.path.join(base, "uploads", "RawData*.xlsx"))
        + glob.glob(os.path.join(base, "data", "RawData*.xlsx"))
    )
    if not candidates:
        print("[WARN] Records table is empty. Upload data via the dashboard or run:")
        print("   python scripts/import_data.py --excel uploads/RawData.xlsx --sheet DataEntry")
        return

    xlsx_path = candidates[0]
    print(f"⚙  Auto-importing from {xlsx_path} ...")
    try:
        import openpyxl
        from app.database import Record

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["DataEntry"] if "DataEntry" in wb.sheetnames else wb.active

        # Detect header row (row 1 or 2)
        row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
        if any(str(h).strip() in ("Zone", "Scheme", "Month", "Year") for h in (row2 or [])):
            headers = [str(c).strip() if c else "" for c in row2]
            data_start = 3
        else:
            headers = [str(c).strip() if c else "" for c in row1]
            data_start = 2

        # Header → Record column mapping
        HMAP = {
            "Zone": "zone", "Scheme": "scheme", "Fiscal Year": "fiscal_year",
            "Year": "year", "Month No.": "month_no", "Month": "month", "Quarter": "quarter",
            "Volume Produced (m³)": "vol_produced",
            "Vol Billed Individual Postpaid": "vol_billed_indiv_pp",
            "Vol Billed CWP Postpaid": "vol_billed_cwp_pp",
            "Vol Billed Institutions Postpaid": "vol_billed_inst_pp",
            "Vol Billed Commercial Postpaid": "vol_billed_comm_pp",
            "TOTAL Vol Billed Postpaid": "total_vol_billed_pp",
            "Vol Billed Individual Prepaid": "vol_billed_indiv_prepaid",
            "Vol Billed CWP Prepaid": "vol_billed_cwp_prepaid",
            "Vol Billed Institutions Prepaid": "vol_billed_inst_prepaid",
            "Vol Billed Commercial Prepaid": "vol_billed_comm_prepaid",
            "TOTAL Vol Billed Prepaid": "total_vol_billed_prepaid",
            "TOTAL Revenue Water m³": "revenue_water", "Non-Revenue Water m³": "nrw",
            "% NRW": "pct_nrw",
            "Chlorine kg": "chlorine_kg", "Alum Sulphate kg": "alum_kg",
            "Soda Ash kg": "soda_ash_kg", "Algae Floc litres": "algae_floc_litres",
            "Sud Floc litres": "sud_floc_litres", "Potassium Permanganate kg": "kmno4_kg",
            "Cost of Chemicals MWK": "chem_cost", "Chem Cost per m³": "chem_cost_per_m3",
            "Power Usage kWh": "power_kwh", "Cost of Power MWK": "power_cost",
            "Power Cost per m³": "power_cost_per_m3",
            "Distances Covered km": "distances_km", "Fuel Used litres": "fuel_used_litres",
            "Cost of Fuel MWK": "fuel_cost", "Maintenance MWK": "maintenance",
            "Staff Costs MWK": "staff_costs", "Wages MWK": "wages",
            "Other Overhead MWK": "other_overhead",
            "TOTAL Operating Costs MWK": "op_cost",
            "OpCost per m³ Produced": "op_cost_per_m3_produced",
            "OpCost per m³ Billed": "op_cost_per_m3_billed",
            "Permanent Staff": "perm_staff", "Temporary Staff": "temp_staff",
            "ALL Conn BroughtFwd": "all_conn_bfwd", "ALL Conn Applied": "all_conn_applied",
            "ALL Conn TOTAL Done": "new_connections", "ALL Conn CarriedFwd": "all_conn_cfwd",
            "Prepaid Meters Installed": "prepaid_meters_installed",
            "Disconnected Individual": "disconnected_individual",
            "Disconnected Institutional": "disconnected_inst",
            "Disconnected Commercial": "disconnected_commercial",
            "Disconnected CWP": "disconnected_cwp", "TOTAL Disconnected": "total_disconnected",
            "Active Postpaid Individual": "active_post_individual",
            "Active Postpaid Institutional": "active_post_inst",
            "Active Postpaid Commercial": "active_post_commercial",
            "Active Postpaid CWP": "active_post_cwp",
            "TOTAL Active Postpaid": "active_postpaid",
            "Active Prepaid Individual": "active_prep_individual",
            "Active Prepaid Institutional": "active_prep_inst",
            "Active Prepaid Commercial": "active_prep_commercial",
            "Active Prepaid CWP": "active_prep_cwp",
            "TOTAL Active Prepaid": "active_prepaid",
            "TOTAL Active Customers": "active_customers",
            "Total Metered Consumers": "total_metered",
            "Population Supply Area": "pop_supply_area",
            "Population Supplied": "pop_supplied",
            "Pct Population Supplied": "pct_pop_supplied",
            "ALL StuckM BroughtFwd": "stuck_meters", "ALL StuckM New": "stuck_new",
            "ALL StuckM Repaired": "stuck_repaired", "ALL StuckM Replaced": "stuck_replaced",
            "TOTAL Pipe Breakdowns": "pipe_breakdowns", "Pump Breakdowns": "pump_breakdowns",
            "Pump Hours Lost": "pump_hours_lost",
            "Normal Supply Hours": "supply_hours", "Power Failure Hours": "power_fail_hours",
            "DevLines 32mm": "dev_lines_32mm", "DevLines 50mm": "dev_lines_50mm",
            "DevLines 63mm": "dev_lines_63mm", "DevLines 90mm": "dev_lines_90mm",
            "DevLines 110mm": "dev_lines_110mm", "TOTAL Dev Lines Done": "dev_lines_total",
            "TOTAL Cash Coll PP": "cash_coll_pp", "TOTAL Cash Coll Prepaid": "cash_coll_prepaid",
            "TOTAL Cash Collected": "cash_collected",
            "TOTAL Amt Billed PP": "amt_billed_pp",
            "TOTAL Amt Billed Prepaid": "amt_billed_prepaid",
            "TOTAL Amount Billed": "amt_billed",
            "TOTAL Service Charge": "service_charge", "TOTAL Meter Rental": "meter_rental",
            "TOTAL Sales MWK": "total_sales",
            "Private Debtors MWK": "private_debtors", "Public Debtors MWK": "public_debtors",
            "TOTAL Debtors MWK": "total_debtors",
            "OpCost per Sales": "op_cost_per_sales",
            "Cash Collection Rate": "collection_rate",
            "Collection per Total Sales": "collection_per_sales",
            "Cust Applied Connection": "conn_applied",
            "Days to Quotation": "days_to_quotation",
            "Cust Fully Paid": "conn_fully_paid", "Days to Connect": "days_to_connect",
            "Connectivity Rate": "connectivity_rate",
            "Queries Received": "queries_received",
            "Time to Resolve Queries": "time_to_resolve",
            "Response Time avg": "response_time_avg",
        }

        rec_cols = {c.name for c in Record.__table__.columns if c.name != "id"}
        str_cols = {"zone", "scheme", "fiscal_year", "month", "quarter"}

        col_map = {}
        for i, h in enumerate(headers):
            if h in HMAP and HMAP[h] in rec_cols:
                col_map[i] = HMAP[h]

        inserted = 0
        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=True):
            data = {}
            for i, db_col in col_map.items():
                v = row[i] if i < len(row) else None
                if db_col in str_cols:
                    data[db_col] = str(v) if v is not None else ""
                else:
                    try:
                        data[db_col] = float(v) if v is not None else 0.0
                    except (ValueError, TypeError):
                        data[db_col] = 0.0
            if not data.get("zone") or data["zone"] in ("Zone", ""):
                continue
            data["year"] = int(data.get("year", 0))
            data["month_no"] = int(data.get("month_no", 0))
            db.add(Record(**data))
            inserted += 1

        db.commit()
        print(f"[OK] Auto-imported {inserted} records from {os.path.basename(xlsx_path)}")
    except Exception as e:
        print(f"[FAIL] Auto-import failed: {e}")
        db.rollback()


app = FastAPI(
    title="SRWB Operations Dashboard API",
    description=(
        "Backend API for the Southern Region Water Board "
        "Operations & Performance Dashboard.\n\n"
        "All monetary values are in **MWK (Malawian Kwacha)**. "
        "Volume in **m³**. Financial year runs April → March.\n\n"
        "**Authentication:** `POST /api/auth/login` with username + password "
        "to obtain a Bearer token.  Include it as:\n"
        "`Authorization: Bearer <token>`\n\n"
        "**Roles:** `admin` · `user` · `viewer`"
    ),
    version="2.0.0",
    contact={"name": "SRWB IT / Corporate Planning"},
    license_info={"name": "Internal Use"},
    lifespan=lifespan,
)

app.state.limiter = _limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SlowAPIMiddleware)

# ── CORS ──────────────────────────────────────────────────────
# Set SRWB_ALLOWED_ORIGINS to a comma-separated list of origins in
# production, e.g. "https://dashboard.srwb.mw,https://ops.srwb.mw"
# Restricted localhost defaults for development; production validation blocks '*'.
_allowed_origins = settings.allowed_origins

app.add_middleware(
    CORSMiddleware,
    allow_origins=_allowed_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE"],
    allow_headers=["*"],
)

# ── Auth endpoints (public — no auth dependency) ───────────────
app.include_router(auth_router)

# ── Admin user-management (admin role required) ───────────────
app.include_router(admin_router, dependencies=[Depends(require_admin)])

# ── Data read endpoints (any authenticated user) ──────────────
# require_export on /export/csv is enforced inside records.py
app.include_router(records.router,   dependencies=[Depends(get_current_user)])
app.include_router(analytics.router, dependencies=[Depends(get_current_user)])
app.include_router(budget.router,    dependencies=[Depends(get_current_user)])
app.include_router(catalogue.router, dependencies=[Depends(get_current_user)])
app.include_router(panels.router,    dependencies=[Depends(get_current_user)])
app.include_router(reports.router,   dependencies=[Depends(get_current_user)])
app.include_router(insights.router,  dependencies=[Depends(get_current_user)])

# ── Upload (admin only) ───────────────────────────────────────
app.include_router(upload.router, dependencies=[Depends(require_admin)])

# ── Static assets ─────────────────────────────────────────────
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
INDEX_PATH  = os.path.join(STATIC_DIR, "index.html")
if os.path.isdir(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.middleware("http")
async def add_request_context(request: Request, call_next):
    request_id = request.headers.get("X-Request-ID") or str(uuid.uuid4())
    token = REQUEST_ID_CTX.set(request_id)
    start = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        duration_ms = round((time.perf_counter() - start) * 1000, 2)
        app_logger.exception(
            "request_failed",
            extra={
                "method": request.method,
                "path": request.url.path,
                "duration_ms": duration_ms,
                "request_id": request_id,
            },
        )
        REQUEST_ID_CTX.reset(token)
        raise
    duration_ms = round((time.perf_counter() - start) * 1000, 2)
    response.headers["X-Request-ID"] = request_id
    app_logger.info(
        "request_completed",
        extra={
            "method": request.method,
            "path": request.url.path,
            "status_code": response.status_code,
            "duration_ms": duration_ms,
            "request_id": request_id,
        },
    )
    REQUEST_ID_CTX.reset(token)
    return response


# ── Root — inject API base URL then serve dashboard ───────────
@app.get("/", include_in_schema=False)
async def serve_dashboard(request: Request):
    if not os.path.exists(INDEX_PATH):
        return {"message": "SRWB API running. Place index.html in app/static/"}
    base_url = str(request.base_url).rstrip("/")
    with open(INDEX_PATH, encoding="utf-8") as f:
        content = f.read()
    content = content.replace("__API_BASE__", base_url)
    return HTMLResponse(
        content=content,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
        }
    )


# ── Health check (public) ─────────────────────────────────────
@app.get("/health", tags=["System"])
def health():
    return {"status": "ok", "version": app.version}


@app.get("/api/debug/db-status", tags=["System"], dependencies=[Depends(require_admin)])
def db_status():
    """Public diagnostic endpoint — shows record counts per year/month."""
    from app.database import Record
    db = SessionLocal()
    try:
        total = db.query(Record).count()
        from sqlalchemy import func
        breakdown = (
            db.query(Record.year, Record.month_no, func.count())
            .group_by(Record.year, Record.month_no)
            .order_by(Record.year, Record.month_no)
            .all()
        )
        sample = db.query(Record).first()
        sample_data = {}
        if sample:
            for col in ("zone", "scheme", "year", "month_no", "month",
                        "vol_produced", "amt_billed", "cash_collected", "op_cost",
                        "active_customers", "total_metered", "stuck_meters", "nrw"):
                sample_data[col] = getattr(sample, col, None)
        return {
            "total_records": total,
            "by_year_month": [{"year": y, "month_no": m, "count": c} for y, m, c in breakdown],
            "sample_row": sample_data,
        }
    finally:
        db.close()
