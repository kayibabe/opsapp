#!/usr/bin/env python3
"""
scripts/import_data.py

Seeds the SQLite database from one of two sources:
  1.  A JSON file produced by extract_from_html.py  (default)
  2.  An Excel file (RawData.xlsx) if --excel is passed

Usage
-----
# From JSON (fastest, use after extract_from_html.py)
python scripts/import_data.py

# From Excel directly (requires openpyxl)
python scripts/import_data.py --excel data/RawData.xlsx --sheet "DataEntry"

Options
-------
--json PATH     Path to JSON file  (default: data/records.json)
--excel PATH    Path to Excel file
--sheet NAME    Sheet name in Excel  (default: first sheet)
--clear         Drop all existing records before importing
"""
import argparse
import json
import sys
import os

# Make sure the project root is on sys.path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import create_tables, SessionLocal, Record

# ── Column mapping: JSON/Excel key → ORM field name ──────────
FIELD_MAP = {
    # --- JSON camelCase keys (from extract_from_html.py) ---
    "zone":            "zone",
    "scheme":          "scheme",
    "month":           "month",
    "monthNo":         "month_no",
    "year":            "year",
    "quarter":         "quarter",
    "volProduced":     "vol_produced",
    "revenueWater":    "revenue_water",
    "nrw":             "nrw",
    "pctNRW":          "pct_nrw",
    "chemCost":        "chem_cost",
    "powerCost":       "power_cost",
    "powerKwh":        "power_kwh",
    "fuelCost":        "fuel_cost",
    "staffCosts":      "staff_costs",
    "opCost":          "op_cost",
    "newConnections":  "new_connections",
    "activeCustomers": "active_customers",
    "activePostpaid":  "active_postpaid",
    "activePrepaid":   "active_prepaid",
    "stuckMeters":     "stuck_meters",
    "stuckNew":        "stuck_new",
    "stuckRepaired":   "stuck_repaired",
    "pipeBreakdowns":  "pipe_breakdowns",
    "pumpBreakdowns":  "pump_breakdowns",
    "supplyHours":     "supply_hours",
    "powerFailHours":  "power_fail_hours",
    "cashCollected":   "cash_collected",
    "amtBilled":       "amt_billed",
    "serviceCharge":   "service_charge",
    "meterRental":     "meter_rental",
    "totalSales":      "total_sales",
    "privateDebtors":  "private_debtors",
    "publicDebtors":   "public_debtors",
    "totalDebtors":    "total_debtors",
    "collectionRate":  "collection_rate",
    "popSupplied":     "pop_supplied",
    "popSupplyArea":   "pop_supply_area",
    "permStaff":       "perm_staff",
    "tempStaff":       "temp_staff",
    # --- Excel column headers (from RawData.xlsx row 2) ---
    "Zone": "zone", "Scheme": "scheme", "Fiscal Year": "fiscal_year",
    "Year": "year", "Month No.": "month_no", "Month": "month", "Quarter": "quarter",
    "Volume Produced (m³)": "vol_produced",
    "Vol Billed Individual Postpaid": "vol_billed_indiv_pp",
    "Vol Billed CWP Postpaid": "vol_billed_cwp_pp",
    "Vol Billed Institutions Postpaid": "vol_billed_inst_pp",
    "Vol Billed Commercial Postpaid": "vol_billed_comm_pp",
    "TOTAL Vol Billed Postpaid": "total_vol_billed_pp",
    "Vol Billed Individual Prepaid": "vol_billed_indiv_prepaid",
    "Vol Billed CWP Prepaid": "vol_billed_cwp_prepaid",
    "Vol Billed Institutions Prepaid": "vol_billed_inst_prepaid",
    "Vol Billed Commercial Prepaid": "vol_billed_comm_prepaid",
    "TOTAL Vol Billed Prepaid": "total_vol_billed_prepaid",
    "TOTAL Revenue Water m³": "revenue_water",
    "Non-Revenue Water m³": "nrw", "% NRW": "pct_nrw",
    "Chlorine kg": "chlorine_kg", "Alum Sulphate kg": "alum_kg",
    "Soda Ash kg": "soda_ash_kg", "Algae Floc litres": "algae_floc_litres",
    "Sud Floc litres": "sud_floc_litres", "Potassium Permanganate kg": "kmno4_kg",
    "Cost of Chemicals MWK": "chem_cost", "Chem Cost per m³": "chem_cost_per_m3",
    "Power Usage kWh": "power_kwh", "Cost of Power MWK": "power_cost",
    "Power Cost per m³": "power_cost_per_m3",
    "Distances Covered km": "distances_km", "Fuel Used litres": "fuel_used_litres",
    "Cost of Fuel MWK": "fuel_cost", "Maintenance MWK": "maintenance",
    "Staff Costs MWK": "staff_costs", "Wages MWK": "wages",
    "Other Overhead MWK": "other_overhead",
    "TOTAL Operating Costs MWK": "op_cost",
    "OpCost per m³ Produced": "op_cost_per_m3_produced",
    "OpCost per m³ Billed": "op_cost_per_m3_billed",
    "Permanent Staff": "perm_staff", "Temporary Staff": "temp_staff",
    "ALL Conn BroughtFwd": "all_conn_bfwd", "ALL Conn Applied": "all_conn_applied",
    "ALL Conn TOTAL Done": "new_connections", "ALL Conn CarriedFwd": "all_conn_cfwd",
    "Prepaid Meters Installed": "prepaid_meters_installed",
    "Disconnected Individual": "disconnected_individual",
    "Disconnected Institutional": "disconnected_inst",
    "Disconnected Commercial": "disconnected_commercial",
    "Disconnected CWP": "disconnected_cwp", "TOTAL Disconnected": "total_disconnected",
    "Active Postpaid Individual": "active_post_individual",
    "Active Postpaid Institutional": "active_post_inst",
    "Active Postpaid Commercial": "active_post_commercial",
    "Active Postpaid CWP": "active_post_cwp",
    "TOTAL Active Postpaid": "active_postpaid",
    "Active Prepaid Individual": "active_prep_individual",
    "Active Prepaid Institutional": "active_prep_inst",
    "Active Prepaid Commercial": "active_prep_commercial",
    "Active Prepaid CWP": "active_prep_cwp",
    "TOTAL Active Prepaid": "active_prepaid",
    "TOTAL Active Customers": "active_customers",
    "Total Metered Consumers": "total_metered",
    "Population Supply Area": "pop_supply_area",
    "Population Supplied": "pop_supplied", "Pct Population Supplied": "pct_pop_supplied",
    "ALL StuckM BroughtFwd": "stuck_meters", "ALL StuckM New": "stuck_new",
    "ALL StuckM Repaired": "stuck_repaired", "ALL StuckM Replaced": "stuck_replaced",
    "TOTAL Pipe Breakdowns": "pipe_breakdowns", "Pump Breakdowns": "pump_breakdowns",
    "PVC 20mm": "pvc_20mm", "PVC 25mm": "pvc_25mm", "PVC 32mm": "pvc_32mm",
    "PVC 40mm": "pvc_40mm", "PVC 50mm": "pvc_50mm", "PVC 63mm": "pvc_63mm",
    "PVC 75mm": "pvc_75mm", "PVC 90mm": "pvc_90mm", "PVC 110mm": "pvc_110mm",
    "PVC 160mm": "pvc_160mm", "PVC 200mm": "pvc_200mm", "PVC 250mm": "pvc_250mm",
    "PVC 315mm": "pvc_315mm",
    "GI 15mm": "_gi_15mm", "GI 20mm": "_gi_20mm", "GI 25mm": "_gi_25mm", "GI 40mm": "_gi_40mm",
    "GI 50mm": "_gi_50mm", "GI 75mm": "_gi_75mm", "GI 100mm": "_gi_100mm", "GI 150mm": "_gi_150mm", "GI 200mm": "_gi_200mm",
    "DI 150mm": "_di_150mm", "DI 200mm": "_di_200mm", "DI 250mm": "_di_250mm", "DI 300mm": "_di_300mm", "DI 350mm": "_di_350mm", "DI 525mm": "_di_525mm",
    "HDPE 20mm": "_hdpe_20mm", "HDPE 25mm": "_hdpe_25mm", "HDPE 32mm": "_hdpe_32mm", "HDPE 50mm": "_hdpe_50mm",
    "AC 50mm": "_ac_50mm", "AC 75mm": "_ac_75mm", "AC 100mm": "_ac_100mm", "AC 150mm": "_ac_150mm",
    "Pump Hours Lost": "pump_hours_lost",
    "Normal Supply Hours": "supply_hours", "Power Failure Hours": "power_fail_hours",
    "DevLines 32mm": "dev_lines_32mm", "DevLines 50mm": "dev_lines_50mm",
    "DevLines 63mm": "dev_lines_63mm", "DevLines 90mm": "dev_lines_90mm",
    "DevLines 110mm": "dev_lines_110mm", "TOTAL Dev Lines Done": "dev_lines_total",
    "TOTAL Cash Coll PP": "cash_coll_pp", "TOTAL Cash Coll Prepaid": "cash_coll_prepaid",
    "TOTAL Cash Collected": "cash_collected",
    "TOTAL Amt Billed PP": "amt_billed_pp", "TOTAL Amt Billed Prepaid": "amt_billed_prepaid",
    "TOTAL Amount Billed": "amt_billed",
    "TOTAL Service Charge": "service_charge", "TOTAL Meter Rental": "meter_rental",
    "TOTAL Sales MWK": "total_sales",
    "Private Debtors MWK": "private_debtors", "Public Debtors MWK": "public_debtors",
    "TOTAL Debtors MWK": "total_debtors",
    "OpCost per Sales": "op_cost_per_sales", "Cash Collection Rate": "collection_rate",
    "Collection per Total Sales": "collection_per_sales",
    "Cust Applied Connection": "conn_applied", "Days to Quotation": "days_to_quotation",
    "Cust Fully Paid": "conn_fully_paid", "Days to Connect": "days_to_connect",
    "Connectivity Rate": "connectivity_rate",
    "Queries Received": "queries_received", "Time to Resolve Queries": "time_to_resolve",
    "Response Time avg": "response_time_avg",
}


def map_row(raw: dict) -> dict:
    """Convert JSON/Excel row to ORM-ready snake_case dict, including breakdown rollups."""
    mapped = {}
    for src, dst in FIELD_MAP.items():
        if src in raw:
            v = raw[src]
            mapped[dst] = float(v) if isinstance(v, (int, float)) and dst not in ("zone","scheme","month","quarter","fiscal_year") else v

    def _sum(fields):
        total = 0.0
        for field in fields:
            v = mapped.get(field)
            if isinstance(v, (int, float)):
                total += float(v)
        return total

    pvc_fields = ['pvc_20mm','pvc_25mm','pvc_32mm','pvc_40mm','pvc_50mm','pvc_63mm','pvc_75mm','pvc_90mm','pvc_110mm','pvc_160mm','pvc_200mm','pvc_250mm','pvc_315mm']
    gi_fields = ['_gi_15mm','_gi_20mm','_gi_25mm','_gi_40mm','_gi_50mm','_gi_75mm','_gi_100mm','_gi_150mm','_gi_200mm']
    di_fields = ['_di_150mm','_di_200mm','_di_250mm','_di_300mm','_di_350mm','_di_525mm']
    hdpe_ac_fields = ['_hdpe_20mm','_hdpe_25mm','_hdpe_32mm','_hdpe_50mm','_ac_50mm','_ac_75mm','_ac_100mm','_ac_150mm']
    mapped['pipe_pvc'] = _sum(pvc_fields)
    mapped['pipe_gi'] = _sum(gi_fields)
    mapped['pipe_di'] = _sum(di_fields)
    mapped['pipe_hdpe_ac'] = _sum(hdpe_ac_fields)
    for field in gi_fields + di_fields + hdpe_ac_fields:
        mapped.pop(field, None)
    return mapped


def import_from_json(path: str) -> list:
    with open(path) as f:
        return json.load(f)


def import_from_excel(path: str, sheet: str | None = None) -> list:
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb.active

    # Try row 1 first; if it looks like section headers, use row 2
    row1 = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    row2 = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=2, max_row=2))]

    # Heuristic: if row 2 contains "Zone" or "Scheme", it's the header row
    if any(h in ("Zone", "Scheme", "Month", "Year") for h in row2):
        headers = row2
        data_start = 3
        print(f"  Detected header row: 2 (row 1 has section labels)")
    else:
        headers = row1
        data_start = 2

    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        rows.append(dict(zip(headers, row)))
    print(f"  Read {len(rows)} rows from '{ws.title}'")
    return rows


def seed(raw_rows: list, clear: bool = False) -> tuple[int, int, list]:
    create_tables()
    db = SessionLocal()
    inserted = skipped = 0
    errors = []

    try:
        if clear:
            deleted = db.query(Record).delete()
            db.commit()
            print(f"  Cleared {deleted} existing records")

        for i, raw in enumerate(raw_rows):
            try:
                data = map_row(raw)
                if not all(k in data for k in ("zone", "scheme", "month", "year")):
                    skipped += 1
                    errors.append(f"Row {i}: missing required identity fields")
                    continue

                # Upsert: skip if already present (use --clear to replace all)
                exists = (
                    db.query(Record)
                    .filter(
                        Record.zone   == data["zone"],
                        Record.scheme == data["scheme"],
                        Record.month  == data["month"],
                        Record.year   == int(data["year"]),
                    )
                    .first()
                )
                if exists:
                    skipped += 1
                    continue

                db.add(Record(**data))
                inserted += 1

                if inserted % 100 == 0:
                    db.commit()

            except Exception as e:
                errors.append(f"Row {i}: {e}")
                skipped += 1

        db.commit()
    finally:
        db.close()

    return inserted, skipped, errors


def main():
    parser = argparse.ArgumentParser(description="Seed SRWB SQLite database")
    parser.add_argument("--json",  default="data/records.json", help="Path to JSON file")
    parser.add_argument("--excel", default=None,                help="Path to Excel file")
    parser.add_argument("--sheet", default=None,                help="Excel sheet name")
    parser.add_argument("--clear", action="store_true",         help="Clear existing records first")
    args = parser.parse_args()

    print("SRWB Data Import")
    print("=" * 40)

    if args.excel:
        print(f"  Source : Excel → {args.excel}")
        raw_rows = import_from_excel(args.excel, args.sheet)
    else:
        json_path = args.json
        if not os.path.exists(json_path):
            print(f"ERROR: JSON file not found: {json_path}")
            print("Run scripts/extract_from_html.py first, or pass --excel <path>")
            sys.exit(1)
        print(f"  Source : JSON → {json_path}")
        raw_rows = import_from_json(json_path)

    print(f"  Rows   : {len(raw_rows)}")
    print(f"  Clear  : {'yes' if args.clear else 'no (skip duplicates)'}")
    print()

    inserted, skipped, errors = seed(raw_rows, clear=args.clear)

    print(f"  ✓ Inserted : {inserted}")
    print(f"  ⊘ Skipped  : {skipped}")
    if errors:
        print(f"  ✗ Errors   : {len(errors)}")
        for e in errors[:10]:
            print(f"    {e}")
    print()
    print("Done.")


if __name__ == "__main__":
    main()
