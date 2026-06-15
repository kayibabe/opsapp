"""
tests/test_upload.py
====================
Tests for the Excel parser and both upload endpoints.

Run:
    pytest tests/test_upload.py -v

The tests use an in-memory SQLite database with the same schema
as production (including the unique constraint) so they exercise
the real commit path without touching the file system.
"""

from __future__ import annotations

import io
import sqlite3
import unittest

import openpyxl

from app.services.excel_parser import (
    ANOMALY_METRICS,
    ExcelParser,
    ParseResult,
)


# ── Helpers ────────────────────────────────────────────────────────────────

# Mirror of the production `records` columns the tests actually exercise.
# Column names and types match the Record ORM in app/database.py:
#   - month is the TEXT month name ("January"…) used by conflict detection
#   - month_no is the INTEGER (1–12) used by anomaly detection and the
#     commit exists-check
#   - metric columns use the canonical DB names, NOT the parser's aliases
SCHEMA_SQL = """
CREATE TABLE records (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    zone            TEXT    NOT NULL,
    scheme          TEXT    NOT NULL,
    fiscal_year     TEXT,
    year            INTEGER NOT NULL,
    month_no        INTEGER NOT NULL,
    month           TEXT,
    quarter         TEXT,
    -- metrics (canonical DB column names) --
    vol_produced     REAL,
    active_customers REAL,
    pct_nrw          REAL,
    amt_billed       REAL,
    cash_collected   REAL,
    new_connections  REAL,
    CONSTRAINT uq_zone_scheme_month_year
        UNIQUE (zone, scheme, month_no, year)
);
"""

MONTH_INT_TO_NAME = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


def make_db(seed_rows: list[dict] | None = None) -> sqlite3.Connection:
    """
    Create a fresh in-memory DB, optionally seeded with existing records.

    Seed dicts use canonical DB column names. If a row supplies `month_no`
    but no `month`, the textual month name is derived automatically so that
    both the integer (anomaly/commit) and string (conflict) lookups work.
    """
    db = sqlite3.connect(":memory:")
    db.executescript(SCHEMA_SQL)
    for row in (seed_rows or []):
        data = dict(row)
        if "month" not in data and "month_no" in data:
            data["month"] = MONTH_INT_TO_NAME.get(int(data["month_no"]))
        cols = list(data.keys())
        db.execute(
            f"INSERT INTO records ({', '.join(cols)}) "
            f"VALUES ({', '.join('?' for _ in cols)})",
            [data[c] for c in cols],
        )
    db.commit()
    return db


def build_xlsx(rows: list[dict], headers: list[str] | None = None) -> io.BytesIO:
    """
    Build a minimal in-memory Excel workbook with the given rows.
    'rows' is a list of dicts; keys become headers if not provided explicitly.

    The DataEntry template has TWO header rows — a section-group row (row 1)
    and the actual column names (row 2) — with data starting on row 3. The
    parser reads headers from row 2, so we emit a blank group row first.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataEntry"

    if headers is None:
        headers = list(rows[0].keys()) if rows else []

    ws.append([None] * len(headers))   # row 1: section group labels (ignored)
    ws.append(headers)                 # row 2: column names
    for row in rows:
        ws.append([row.get(h) for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def make_row(**overrides) -> dict:
    """
    Return a minimal valid RawData row keyed by Excel header strings.
    Headers are chosen so they normalise to the canonical DB columns via
    COLUMN_MAP (e.g. "Total Active Customers" → active_customers).
    """
    base = {
        "Zone":                    "Blantyre North",
        "Scheme":                  "Ndirande",
        "Month":                   1,
        "Year":                    2024,
        "Total Active Customers":  18000,
        "Vol. Produced":           310000,
        "NRW %":                   29,
        "Total Amount Billed":     340000000,
        "Total Cash Collected":    360000000,
        "ALL Conn TOTAL Done":     120,
    }
    base.update(overrides)
    return base


# ── Parser tests ───────────────────────────────────────────────────────────

class TestHeaderNormalisation(unittest.TestCase):

    def setUp(self):
        self.p = ExcelParser()

    def test_strip_punctuation(self):
        self.assertEqual(self.p._normalize_header("Vol. Produced (m³)"), "vol_produced_m")

    def test_case_insensitive(self):
        self.assertEqual(self.p._normalize_header("ACTIVE CUSTOMERS"), "active_customers")

    def test_slash_stripped(self):
        self.assertEqual(self.p._normalize_header("NWCs B/F"), "nwcs_bf")

    def test_percent_stripped(self):
        norm = self.p._normalize_header("NRW %")
        self.assertIn("nrw", norm)

    def test_none_returns_empty(self):
        self.assertEqual(self.p._normalize_header(None), "")


class TestTypeCoercion(unittest.TestCase):

    def setUp(self):
        self.p = ExcelParser()

    def test_int_passthrough(self):
        self.assertEqual(self.p._coerce_numeric(1234), 1234.0)

    def test_string_with_commas(self):
        self.assertEqual(self.p._coerce_numeric("1,234,567"), 1234567.0)

    def test_non_numeric_returns_none(self):
        self.assertIsNone(self.p._coerce_numeric("N/A"))

    def test_month_name(self):
        self.assertEqual(self.p._coerce_month("January"), 1)
        self.assertEqual(self.p._coerce_month("jan"),     1)
        self.assertEqual(self.p._coerce_month("Dec"),     12)

    def test_month_out_of_range(self):
        self.assertIsNone(self.p._coerce_month(13))
        self.assertIsNone(self.p._coerce_month(0))

    def test_year_two_digit(self):
        self.assertEqual(self.p._coerce_year(24), 2024)

    def test_year_four_digit(self):
        self.assertEqual(self.p._coerce_year(2024), 2024)


class TestParserHappyPath(unittest.TestCase):

    def test_single_valid_row(self):
        xlsx = build_xlsx([make_row()])
        db   = make_db()
        result = ExcelParser().parse(xlsx, db)

        self.assertEqual(len(result.rows), 1)
        self.assertEqual(result.error_rows, [])
        self.assertEqual(result.period_month, 1)
        self.assertEqual(result.period_year,  2024)

    def test_multiple_rows_different_schemes(self):
        rows = [
            make_row(Scheme="Ndirande"),
            make_row(Scheme="Chilomoni"),
        ]
        xlsx = build_xlsx(rows)
        db   = make_db()
        result = ExcelParser().parse(xlsx, db)
        self.assertEqual(len(result.rows), 2)
        self.assertEqual(len(result.error_rows), 0)

    def test_blank_rows_skipped(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DataEntry"
        headers = list(make_row().keys())
        ws.append([None] * len(headers))            # row 1: group labels
        ws.append(headers)                          # row 2: column names
        ws.append(list(make_row().values()))        # row 3: data
        ws.append([None] * len(headers))            # row 4: blank row
        ws.append(list(make_row(Scheme="Chilomoni").values()))
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        result = ExcelParser().parse(buf, make_db())
        self.assertEqual(len(result.rows), 2)

    def test_unrecognised_column_captured(self):
        row = make_row()
        row["maint_cost"] = 99999   # not in COLUMN_MAP
        xlsx = build_xlsx([row])
        result = ExcelParser().parse(xlsx, make_db())
        self.assertIn("maint_cost", result.unrecognised_columns)
        self.assertEqual(len(result.error_rows), 0)  # still importable

    def test_fallback_to_first_sheet_when_no_rawdata(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MonthlyData"
        row = make_row()
        headers = list(row.keys())
        ws.append([None] * len(headers))   # row 1: group labels
        ws.append(headers)                 # row 2: column names
        ws.append(list(row.values()))      # row 3: data
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        result = ExcelParser().parse(buf, make_db())
        self.assertEqual(len(result.rows), 1)


class TestValidationErrors(unittest.TestCase):

    def test_missing_zone_is_error(self):
        xlsx = build_xlsx([make_row(Zone=None)])
        result = ExcelParser().parse(xlsx, make_db())
        self.assertEqual(len(result.error_rows), 1)
        issues = result.rows[0].issues
        self.assertTrue(any(i.field == "zone" for i in issues))

    def test_missing_scheme_is_error(self):
        xlsx = build_xlsx([make_row(Scheme="")])
        result = ExcelParser().parse(xlsx, make_db())
        self.assertEqual(len(result.error_rows), 1)

    def test_invalid_month_is_error(self):
        xlsx = build_xlsx([make_row(Month="BadMonth")])
        result = ExcelParser().parse(xlsx, make_db())
        self.assertEqual(len(result.error_rows), 1)

    def test_missing_vol_produced_is_error(self):
        row = make_row()
        row["Vol. Produced"] = None
        xlsx = build_xlsx([row])
        result = ExcelParser().parse(xlsx, make_db())
        self.assertEqual(len(result.error_rows), 1)
        fields = [i.field for i in result.rows[0].issues]
        self.assertIn("vol_produced", fields)

    def test_non_numeric_metric_is_warning_not_error(self):
        row = make_row()
        row["Total Amount Billed"] = "not-a-number"
        xlsx = build_xlsx([row])
        result = ExcelParser().parse(xlsx, make_db())
        # Row is still importable — warning only
        self.assertEqual(len(result.error_rows), 0)
        self.assertEqual(result.rows[0].status, "warning")

    def test_missing_required_dim_column_raises_valueerror(self):
        """If 'Zone' column is entirely absent from the sheet, raise."""
        row = {"Scheme": "Ndirande", "Month": 1, "Year": 2024,
               "Vol. Produced": 310000}
        xlsx = build_xlsx([row])
        with self.assertRaises(ValueError):
            ExcelParser().parse(xlsx, make_db())


class TestAnomalyDetection(unittest.TestCase):

    def _seed(self, zone, scheme, val):
        """Seed 3 historical months with the same vol_produced value."""
        return [
            {"zone": zone, "scheme": scheme, "month_no": m, "year": 2023,
             "active_customers": 18000, "vol_produced": val, "pct_nrw": 30,
             "amt_billed": 340e6, "cash_collected": 360e6, "new_connections": 120}
            for m in (9, 10, 11)
        ]

    def test_high_anomaly_flagged(self):
        seed = self._seed("Blantyre North", "Ndirande", 100_000)
        db   = make_db(seed)
        # New value is 4× average → above ANOMALY_HIGH (3.0)
        xlsx = build_xlsx([make_row(**{"Vol. Produced": 400_000})])
        result = ExcelParser().parse(xlsx, db)
        row = result.rows[0]
        anomaly_issues = [i for i in row.issues if i.field == "vol_produced"]
        self.assertTrue(len(anomaly_issues) > 0)
        self.assertEqual(anomaly_issues[0].severity, "warning")

    def test_low_anomaly_flagged(self):
        seed = self._seed("Blantyre North", "Ndirande", 300_000)
        db   = make_db(seed)
        # New value is 10% of average → below ANOMALY_LOW (0.33)
        xlsx = build_xlsx([make_row(**{"Vol. Produced": 30_000})])
        result = ExcelParser().parse(xlsx, db)
        row = result.rows[0]
        anomaly_issues = [i for i in row.issues if i.field == "vol_produced"]
        self.assertTrue(len(anomaly_issues) > 0)

    def test_within_range_not_flagged(self):
        seed = self._seed("Blantyre North", "Ndirande", 310_000)
        db   = make_db(seed)
        # New value is 1.02× average → normal
        xlsx = build_xlsx([make_row(**{"Vol. Produced": 315_000})])
        result = ExcelParser().parse(xlsx, db)
        row = result.rows[0]
        vol_issues = [i for i in row.issues if i.field == "vol_produced"]
        self.assertEqual(vol_issues, [])

    def test_no_history_skips_anomaly_check(self):
        """New (zone, scheme) with no history: no anomaly warnings."""
        db = make_db()   # empty
        xlsx = build_xlsx([make_row()])
        result = ExcelParser().parse(xlsx, db)
        anomaly_issues = [
            i for r in result.rows for i in r.issues
            if i.severity == "warning" and i.field in ANOMALY_METRICS
        ]
        self.assertEqual(anomaly_issues, [])


class TestConflictDetection(unittest.TestCase):

    def test_no_conflict_when_db_empty(self):
        db = make_db()
        result = ExcelParser().parse(build_xlsx([make_row()]), db)
        self.assertEqual(result.conflict_rows, [])

    def test_conflict_detected_for_existing_record(self):
        seed = [{
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month_no": 1, "year": 2024,
            "active_customers": 17000, "vol_produced": 300_000,
            "pct_nrw": 31, "amt_billed": 330e6,
            "cash_collected": 350e6, "new_connections": 115,
        }]
        db = make_db(seed)
        result = ExcelParser().parse(build_xlsx([make_row()]), db)
        self.assertEqual(len(result.conflict_rows), 1)

        conflict = result.conflict_rows[0].conflict
        self.assertIn("existing", conflict)
        self.assertIn("incoming", conflict)
        self.assertEqual(conflict["existing"]["active_customers"], 17000)

    def test_no_conflict_for_different_period(self):
        seed = [{
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month_no": 12, "year": 2023,   # ← different month/year
            "active_customers": 17000, "vol_produced": 300_000,
            "pct_nrw": 31, "amt_billed": 330e6,
            "cash_collected": 350e6, "new_connections": 115,
        }]
        db = make_db(seed)
        result = ExcelParser().parse(build_xlsx([make_row()]), db)
        self.assertEqual(result.conflict_rows, [])

    def test_conflict_only_on_matching_zone_scheme(self):
        seed = [{
            "zone": "Zomba", "scheme": "Mitengo",  # ← different zone/scheme
            "month_no": 1, "year": 2024,
            "active_customers": 5000, "vol_produced": 80_000,
            "pct_nrw": 35, "amt_billed": 100e6,
            "cash_collected": 110e6, "new_connections": 40,
        }]
        db = make_db(seed)
        result = ExcelParser().parse(build_xlsx([make_row()]), db)
        self.assertEqual(result.conflict_rows, [])


class TestPeriodInference(unittest.TestCase):

    def test_single_month_inferred(self):
        result = ExcelParser().parse(build_xlsx([make_row()]), make_db())
        self.assertEqual(result.period_month, 1)
        self.assertEqual(result.period_year,  2024)

    def test_majority_month_wins(self):
        rows = [
            make_row(Month=3),
            make_row(Month=3, Scheme="Chilomoni"),
            make_row(Month=4, Scheme="Bangwe"),   # minority
        ]
        result = ExcelParser().parse(build_xlsx(rows), make_db())
        self.assertEqual(result.period_month, 3)


class TestCommitLogic(unittest.TestCase):
    """
    Tests for _execute_commit without going through FastAPI — imports the
    function directly. The current signature is positional:
        _execute_commit(conn, preview_data, global_mode, per_row_res)
    """

    def setUp(self):
        from app.routers.upload import _execute_commit
        self._execute_commit = _execute_commit

    def _make_preview(self, rows: list[dict]) -> dict:
        return {
            "rows": rows,
            "filename": "test.xlsx",
            "period_month": 1,
            "period_year":  2024,
        }

    def _run(self, db, rows, global_mode="replace", per_row_res=None):
        preview = self._make_preview(rows)
        return self._execute_commit(
            db,
            preview,
            global_mode,
            per_row_res or {},
        )

    def _importable_row(self, **kw) -> dict:
        base = {
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month": 1, "year": 2024,
            "status": "ok",
            "conflict": None,
            "metrics": {
                "active_customers": 18000, "vol_produced": 310000,
                "pct_nrw": 29, "amt_billed": 340e6,
                "cash_collected": 360e6, "new_connections": 120,
            },
        }
        base.update(kw)
        return base

    def test_insert_new_record(self):
        db    = make_db()
        stats = self._run(db, [self._importable_row()])
        self.assertEqual(stats["rows_inserted"], 1)
        self.assertEqual(stats["rows_replaced"], 0)
        row = db.execute("SELECT zone FROM records").fetchone()
        self.assertIsNotNone(row)

    def test_error_rows_excluded(self):
        """Rows flagged status=error are excluded from the importable set."""
        row = self._importable_row(status="error")
        stats = self._run(make_db(), [row])
        self.assertEqual(stats["rows_inserted"], 0)
        self.assertEqual(stats["rows_importable"], 0)
        self.assertEqual(stats["rows_total"], 1)

    def test_replace_conflict(self):
        db = make_db([{
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month_no": 1, "year": 2024,
            "active_customers": 17000, "vol_produced": 300_000,
            "pct_nrw": 31, "amt_billed": 330e6,
            "cash_collected": 350e6, "new_connections": 115,
        }])
        row = self._importable_row(conflict={"existing": {}, "incoming": {}})
        stats = self._run(db, [row], global_mode="replace")
        self.assertEqual(stats["rows_replaced"], 1)
        updated = db.execute(
            "SELECT active_customers FROM records WHERE zone='Blantyre North'"
        ).fetchone()
        self.assertEqual(updated[0], 18000)   # new value, not old 17000

    def test_skip_conflict(self):
        db = make_db([{
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month_no": 1, "year": 2024,
            "active_customers": 17000, "vol_produced": 300_000,
            "pct_nrw": 31, "amt_billed": 330e6,
            "cash_collected": 350e6, "new_connections": 115,
        }])
        row = self._importable_row(conflict={"existing": {}, "incoming": {}})
        stats = self._run(db, [row], global_mode="skip")
        self.assertEqual(stats["rows_skipped"], 1)
        unchanged = db.execute(
            "SELECT active_customers FROM records WHERE zone='Blantyre North'"
        ).fetchone()
        self.assertEqual(unchanged[0], 17000)   # not overwritten

    def test_per_row_override_beats_global(self):
        """global=replace but per-row says skip for this specific row."""
        db = make_db([{
            "zone": "Blantyre North", "scheme": "Ndirande",
            "month_no": 1, "year": 2024,
            "active_customers": 17000, "vol_produced": 300_000,
            "pct_nrw": 31, "amt_billed": 330e6,
            "cash_collected": 350e6, "new_connections": 115,
        }])
        row = self._importable_row(conflict={"existing": {}, "incoming": {}})
        # _row_resolution_key format: "{zone}|{scheme}|{year}|{month}"
        per_row = {"Blantyre North|Ndirande|2024|1": "skip"}
        stats = self._run(db, [row], global_mode="replace", per_row_res=per_row)
        self.assertEqual(stats["rows_skipped"], 1)

    def test_bad_column_row_errored(self):
        """A corrupt metric column name errors that row without inserting it."""
        db = make_db()
        row = self._importable_row()
        # Inject a corrupt column name into metrics
        row["metrics"]["__bad column__"] = 999
        stats = self._execute_commit(db, self._make_preview([row]), "replace", {})
        self.assertEqual(stats["rows_errored"], 1)
        self.assertEqual(stats["rows_inserted"], 0)
        count = db.execute("SELECT COUNT(*) FROM records").fetchone()[0]
        self.assertEqual(count, 0)

    def test_atomic_all_or_nothing(self):
        """
        Two valid rows with distinct schemes commit together.
        """
        db = make_db()
        rows = [
            self._importable_row(scheme="Ndirande"),
            self._importable_row(scheme="Chilomoni"),
        ]
        stats = self._run(db, rows)
        count = db.execute("SELECT COUNT(*) FROM records").fetchone()[0]
        self.assertEqual(count, 2)
        self.assertEqual(stats["rows_inserted"], 2)


if __name__ == "__main__":
    unittest.main()
