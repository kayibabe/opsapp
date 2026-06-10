#!/usr/bin/env python3
"""
OpsApp / dataupdater  –  update_rawdata_master.py

Updates RawData.xlsx from 5 source zone workbooks.

HOW THE SMART-DIFF WORKS
─────────────────────────
Before touching any cell the script inspects the master for every
(Zone, Scheme, Month No.) key:

  • Row has data already  →  skipped entirely
  • Row exists but is empty/zero  →  filled from the source file
  • Row is absent from master  →  appended (if ALLOW_APPEND_NEW_ROWS=True)

"Has data" is decided by checking the probe column ('Volume Produced (m³)')
in the master.  Zero is treated as empty when TREAT_ZERO_AS_EMPTY=True.

The script is safe to run at any time – it will never overwrite records
that are already complete, and will fill in only what is still missing.

CLI
───
  python update_rawdata_master.py           # normal run
  python update_rawdata_master.py --test    # dry-run, nothing saved
  python update_rawdata_master.py --force   # overwrite existing values too

Dependencies:  pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import os
import sys
import traceback
from dataclasses import dataclass, field
from typing import Any, Dict, List, Set, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook

# =============================================================================
# CONFIGURATION
# =============================================================================

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

def _p(filename):
    return os.path.join(_SCRIPT_DIR, filename)

RAWDATA_MASTER_FILE      = _p("RawData.xlsx")
OUTPUT_FILE              = _p("RawData_updated.xlsx")  # set to None to overwrite master

ZONE_FILES = {
    "Liwonde":  _p("Liwonde.xlsx"),
    "Mangochi": _p("Mangochi.xlsx"),
    "Mulanje":  _p("Mulanje.xlsx"),
    "Ngabu":    _p("Ngabu.xlsx"),
    "Zomba":    _p("Zomba.xlsx"),
}

# Behaviour flags
TEST_MODE                 = False   # True = preview only, nothing saved
ALLOW_APPEND_NEW_ROWS     = True    # append row when key absent from master
OVERWRITE_EXISTING_VALUES = False   # True = full refresh (replace all values)
TREAT_ZERO_AS_EMPTY       = True    # treat 0 as "missing" in master
CREATE_MASTER_IF_MISSING  = False

# Master sheet
MASTER_SHEET_NAME = "DataEntry"
MASTER_HEADER_ROW = 2
FIRST_DATA_ROW    = 3
KEY_COLUMNS       = ["Zone", "Scheme", "Month No."]

# Probe: row 8 in source = 'Volume Produced (m³)' in master
SOURCE_PROBE_ROW = 8
PROBE_COL_MASTER = "Volume Produced (m³)"

NON_DATA_SHEETS = {
    "Zone Monthly", "1st QTR", "2nd QTR ", "Midyear Summary",
    "3rd QTR", "4th QTR", "Zone Annual Summary", "Targets", "Budget",
    "ZA TH JA CALCULATION", "COST OF POWER", "WATER TARIFF",
}

MONTH_SOURCE_COLUMN = {
    "April": 3, "May": 4, "June": 5, "July": 6, "August": 7,
    "September": 8, "October": 9, "November": 10, "December": 11,
    "January": 12, "February": 13, "March": 14,
}

MONTH_NAME_TO_NUMBER = {
    "April": 4, "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
    "January": 1, "February": 2, "March": 3,
}

MINIMAL_MASTER_COLUMNS = [
    "Zone", "Scheme", "Fiscal Year", "Year", "Month No.", "Month", "Quarter",
    "Volume Produced (m³)",
]

SOURCE_TO_RAWDATA_COL_MAP = [
    (8, 7), (10, 8), (11, 9), (12, 10), (13, 11),
    (16, 13), (17, 14), (18, 15), (19, 16),
    (23, 19), (28, 21), (30, 23), (32, 25), (34, 27), (36, 29), (38, 31),
    (40, 33), (43, 35), (45, 37),
    (48, 39), (49, 40), (50, 41), (51, 42), (53, 43), (54, 44), (55, 45),
    (60, 49), (62, 51),
    (65, 52), (66, 53), (67, 54), (68, 55),
    (71, 58), (72, 59), (73, 60), (74, 61),
    (77, 64), (78, 65), (79, 66), (80, 67),
    (83, 70), (84, 71), (85, 72), (86, 73),
    (97, 80),
    (100, 81), (101, 82), (102, 83), (103, 84),
    (106, 86), (107, 87), (108, 88), (109, 89),
    (112, 91), (113, 92), (114, 93), (115, 94),
    (118, 97),
    (120, 98), (121, 99),
    (124, 101), (125, 102), (126, 103), (127, 104),
    (129, 106), (130, 107), (131, 108), (132, 109),
    (134, 111), (135, 112), (136, 113), (137, 114),
    (139, 116), (140, 117), (141, 118), (142, 119),
    (152, 126), (153, 127), (154, 128), (155, 129), (156, 130), (157, 131),
    (158, 132), (159, 133), (160, 134), (161, 135), (162, 136), (163, 137), (164, 138),
    (166, 139), (167, 140), (168, 141), (169, 142), (170, 143), (171, 144),
    (172, 145), (173, 146), (174, 147),
    (176, 148), (177, 149), (178, 150), (179, 151), (180, 152), (181, 153),
    (183, 154), (184, 155), (185, 156), (186, 157),
    (188, 158), (189, 159), (190, 160), (191, 161),
    (193, 163), (194, 164), (195, 165), (196, 166),
    (198, 167), (199, 168), (200, 169), (201, 170), (202, 171),
    (205, 173), (206, 174), (207, 175), (208, 176),
    (211, 178), (212, 179), (213, 180), (214, 181),
    (218, 184), (219, 185), (220, 186), (221, 187),
    (224, 189), (225, 190), (226, 191), (227, 192),
    (231, 195), (232, 196), (233, 197), (234, 198),
    (237, 200), (238, 201), (239, 202), (240, 203),
    (244, 206), (245, 207),
    (252, 212), (253, 213), (254, 214), (255, 215), (256, 216),
    (260, 219), (261, 220),
]

# =============================================================================
# HELPERS
# =============================================================================

@dataclass
class Summary:
    source_files_read: int = 0
    source_sheets_processed: int = 0
    source_records_seen: int = 0
    skipped_already_complete: int = 0
    rows_updated: int = 0
    rows_added: int = 0
    rows_unchanged: int = 0
    cells_written: int = 0
    skipped_empty_source: int = 0
    skipped_layout_error: int = 0
    skipped_append_disabled: int = 0
    duplicate_keys_master: int = 0
    duplicate_keys_source: int = 0
    warnings: int = 0
    errors: int = 0
    filled_records: List[Tuple] = field(default_factory=list)


# When _CAPTURE is True, every log() line is also appended to _LOG_LINES so a
# caller (e.g. the web Upload tool) can return the run transcript to the UI.
_LOG_LINES: List[str] = []
_CAPTURE: bool = False


def log(msg):
    if _CAPTURE:
        _LOG_LINES.append(str(msg))
    print(msg, flush=True)

def norm_text(v):
    return "" if v is None else str(v).strip()

def norm_num(v):
    if pd.isna(v): return 0.0
    if isinstance(v, str):
        v = v.strip()
        if v == "": return 0.0
    try: return float(v)
    except (TypeError, ValueError) as e: raise ValueError(f"Cannot convert {v!r}") from e

def is_empty(v):
    if v is None: return True
    if isinstance(v, str) and v.strip() == "": return True
    if TREAT_ZERO_AS_EMPTY and v in (0, 0.0): return True
    return False

def quarter(m):
    if m in (4,5,6): return "Q1"
    if m in (7,8,9): return "Q2"
    if m in (10,11,12): return "Q3"
    return "Q4"

def norm_year(v):
    """Coerce a master 'Year' cell to int, or None when blank/unparseable."""
    if v is None:
        return None
    if isinstance(v, str) and v.strip() == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None

def year_for_month(start_year, month_no):
    """SRWB fiscal year runs April->March: Apr-Dec belong to ``start_year``,
    Jan-Mar to ``start_year + 1``."""
    return start_year if month_no >= 4 else start_year + 1

def fiscal_year_label(start_year):
    """2025 -> 'FY2025/26' (matches the format already stored in the master)."""
    return f"FY{start_year}/{str(start_year + 1)[-2:]}"

def parse_fiscal_start_year(fy):
    """Pull the start year out of a fiscal-year label.
    'FY2025/26' -> 2025 ; '2025/26' -> 2025 ; '2025' -> 2025 ; junk -> None."""
    import re
    m = re.search(r"(\d{4})", str(fy or ""))
    return int(m.group(1)) if m else None

def infer_start_year_from_master(master_df):
    """Most common fiscal-year start year already present in the master,
    derived from the (Year, Month No.) pair so it does not depend on the
    'Fiscal Year' string format. Returns None when the master has no usable
    year data."""
    from collections import Counter
    starts = []
    for _, r in master_df.iterrows():
        y = norm_year(r.get("Year"))
        if y is None:
            continue
        try:
            mn = int(r["Month No."])
        except (TypeError, ValueError, KeyError):
            continue
        starts.append(y if mn >= 4 else y - 1)
    if not starts:
        return None
    return Counter(starts).most_common(1)[0][0]

def current_start_year():
    """Fiscal-year start year for today's date (April->March cycle)."""
    import datetime
    today = datetime.date.today()
    return today.year if today.month >= 4 else today.year - 1

def ensure(path, label):
    if not os.path.exists(path):
        raise FileNotFoundError(f"{label} not found: {path}")


def read_master_df(path):
    try:
        return pd.read_excel(path, sheet_name=MASTER_SHEET_NAME, header=MASTER_HEADER_ROW - 1)
    except ValueError as e:
        raise ValueError(f"Expected sheet '{MASTER_SHEET_NAME}' in {path}") from e


def build_index(master_df, summary, fallback_start_year):
    """Index master rows by (Zone, Scheme, Year, Month No.).

    Including the year in the key is what lets a brand-new fiscal year be
    appended instead of being skipped as "already complete" because the same
    (Zone, Scheme, Month) existed in a prior year. When a master row has no
    Year cell, the year is derived from ``fallback_start_year`` so legacy
    masters still index cleanly.
    """
    for col in KEY_COLUMNS:
        if col not in master_df.columns:
            raise KeyError(f"Master missing key column: {col}")
    index = {}
    for i, row in master_df.iterrows():
        month_no = int(row["Month No."])
        year = norm_year(row.get("Year"))
        if year is None:
            year = year_for_month(fallback_start_year, month_no)
        key = (norm_text(row["Zone"]), norm_text(row["Scheme"]), year, month_no)
        excel_row = i + FIRST_DATA_ROW
        if key in index:
            summary.duplicate_keys_master += 1
            log(f"  WARNING: Duplicate master key at row {excel_row}: {key}")
            continue
        index[key] = excel_row
    return index


def build_filled_set(master_index, ws, master_columns, force_overwrite):
    """Keys whose probe column already has a non-empty value in the master."""
    if force_overwrite:
        return set()
    try:
        probe_col_excel = master_columns.index(PROBE_COL_MASTER) + 1
    except ValueError:
        log(f"  WARNING: Probe column '{PROBE_COL_MASTER}' not found – all rows treated as incomplete")
        return set()
    filled = set()
    for key, excel_row in master_index.items():
        val = ws.cell(row=excel_row, column=probe_col_excel).value
        if not is_empty(val):
            filled.add(key)
    return filled


def build_new_row(zone, scheme, month_name, source_df, master_columns,
                  start_year, fy_label, summary=None):
    month_no = MONTH_NAME_TO_NUMBER[month_name]
    row = {c: None for c in master_columns}
    # Year / Fiscal Year are required by the Step-2 importer; without them the
    # appended record fails validation and is silently dropped on import.
    row.update({"Zone": zone, "Scheme": scheme,
                "Fiscal Year": fy_label, "Year": year_for_month(start_year, month_no),
                "Month No.": month_no,
                "Month": month_name, "Quarter": quarter(month_no)})
    src_col = MONTH_SOURCE_COLUMN[month_name]
    for sr, dc in SOURCE_TO_RAWDATA_COL_MAP:
        if dc < len(master_columns):
            # Mirror the fill path: a single unparseable source cell must not
            # abort the whole build — record the error and leave that cell blank.
            try:
                row[master_columns[dc]] = norm_num(source_df.iloc[sr, src_col])
            except Exception as e:
                if summary is not None:
                    summary.errors += 1
                log(f"  ERROR type at '{scheme}'/{month_name}/row {sr}: {e}")
    return row


# =============================================================================
# MAIN
# =============================================================================

def run_update(test_mode=TEST_MODE, force_overwrite=False, fiscal_year=None):
    """Programmatic entry point.

    Runs the smart-diff update and returns a tuple:
        (exit_code: int, summary: Summary, log_lines: List[str])

    Behaves identically to the CLI but never calls sys.exit, so it is safe
    to import and call from the web Upload tool.

    ``fiscal_year`` selects which fiscal year the zone workbooks belong to
    (e.g. ``"FY2026/27"`` or ``"2026"``). The source files carry no year of
    their own, so this is how a brand-new fiscal year is loaded. When omitted
    it defaults to the fiscal year already present in the master (i.e. the
    normal monthly top-up), falling back to the current calendar fiscal year.
    """
    global _CAPTURE, _LOG_LINES
    _CAPTURE = True
    _LOG_LINES = []
    s = Summary()

    try:
        log("=" * 72)
        log("OpsApp DataUpdater  –  RawData Smart-Diff Update")
        log("=" * 72)
        log(f"Mode             : {'TEST / DRY-RUN' if test_mode else 'PRODUCTION'}")
        log(f"Overwrite exist  : {'YES (--force)' if force_overwrite else 'NO  – fill missing/zero only'}")
        log(f"Master file      : {RAWDATA_MASTER_FILE}")
        log(f"Output file      : {OUTPUT_FILE or '(overwrite master)'}")
        log("")

        # ── Validate files ────────────────────────────────────────────────────
        if not os.path.exists(RAWDATA_MASTER_FILE):
            if CREATE_MASTER_IF_MISSING:
                log("Creating minimal master workbook...")
                wb0 = Workbook(); ws0 = wb0.active; ws0.title = MASTER_SHEET_NAME
                for ci, n in enumerate(MINIMAL_MASTER_COLUMNS, 1):
                    ws0.cell(row=MASTER_HEADER_ROW, column=ci).value = n
                wb0.save(RAWDATA_MASTER_FILE)
            else:
                raise FileNotFoundError(f"Master not found: {RAWDATA_MASTER_FILE}")

        ensure(RAWDATA_MASTER_FILE, "Master RawData file")
        for zone, path in ZONE_FILES.items():
            ensure(path, f"Source file for '{zone}'")

        # ── Load master ───────────────────────────────────────────────────────
        log("Loading master workbook...")
        master_df = read_master_df(RAWDATA_MASTER_FILE)
        wb = load_workbook(RAWDATA_MASTER_FILE)
        if MASTER_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Sheet '{MASTER_SHEET_NAME}' not found in master workbook")
        ws = wb[MASTER_SHEET_NAME]
        master_columns = list(master_df.columns)
        log(f"  {len(master_df)} rows  |  {len(master_columns)} columns")

        # ── Resolve the target fiscal year for the source workbooks ──────────
        target_start_year = parse_fiscal_start_year(fiscal_year)
        if target_start_year is None:
            if fiscal_year:
                log(f"  WARNING: Could not parse fiscal year '{fiscal_year}' – auto-detecting instead")
            target_start_year = infer_start_year_from_master(master_df)
        if target_start_year is None:
            target_start_year = current_start_year()
        target_fy_label = fiscal_year_label(target_start_year)
        log(f"Target fiscal yr : {target_fy_label}  (Apr {target_start_year} - Mar {target_start_year + 1})")
        log("")

        log("Building key index...")
        master_index = build_index(master_df, s, target_start_year)

        log("Scanning for already-complete records...")
        filled_keys = build_filled_set(master_index, ws, master_columns, force_overwrite)
        missing_count = len(master_index) - len(filled_keys)
        log(f"  Complete (will skip)  : {len(filled_keys)}")
        log(f"  Incomplete (to fill)  : {missing_count}")
        log("")

        # ── Process source files ──────────────────────────────────────────────
        for zone_name, source_path in ZONE_FILES.items():
            log(f"Reading: {os.path.basename(source_path)}")
            try:
                xls = pd.ExcelFile(source_path)
            except Exception as e:
                s.errors += 1; log(f"  ERROR opening: {e}"); continue

            s.source_files_read += 1
            data_sheets = [sh for sh in xls.sheet_names if sh not in NON_DATA_SHEETS]
            log(f"  {len(data_sheets)} data sheet(s)")

            for sheet_name in data_sheets:
                try:
                    src_df = pd.read_excel(source_path, sheet_name=sheet_name, header=None)
                except Exception as e:
                    s.errors += 1; log(f"  ERROR reading '{sheet_name}': {e}"); continue

                s.source_sheets_processed += 1
                seen = set()

                for month_name, src_col in MONTH_SOURCE_COLUMN.items():
                    s.source_records_seen += 1
                    month_no = MONTH_NAME_TO_NUMBER[month_name]
                    row_year = year_for_month(target_start_year, month_no)
                    key = (zone_name, sheet_name, row_year, month_no)

                    if key in seen:
                        s.duplicate_keys_source += 1; continue
                    seen.add(key)

                    # Check source probe cell
                    try:
                        probe = src_df.iloc[SOURCE_PROBE_ROW, src_col]
                    except Exception as e:
                        s.skipped_layout_error += 1
                        log(f"  WARNING layout at '{sheet_name}'/{month_name}: {e}"); continue

                    if pd.isna(probe) or probe in ("", 0, 0.0):
                        s.skipped_empty_source += 1; continue

                    # ── DIFF check ────────────────────────────────────────
                    if key in filled_keys:
                        s.skipped_already_complete += 1
                        continue  # master already has data – skip

                    # ── Update existing empty row ─────────────────────────
                    if key in master_index:
                        excel_row = master_index[key]
                        written = 0
                        for sr, dc in SOURCE_TO_RAWDATA_COL_MAP:
                            try:
                                new_val = norm_num(src_df.iloc[sr, src_col])
                            except Exception as e:
                                s.errors += 1
                                log(f"  ERROR type at '{sheet_name}'/{month_name}/row {sr}: {e}"); continue

                            old_val = ws.cell(row=excel_row, column=dc + 1).value
                            if (force_overwrite or is_empty(old_val)) and not is_empty(new_val):
                                ws.cell(row=excel_row, column=dc + 1).value = new_val
                                written += 1

                        if written:
                            s.rows_updated += 1; s.cells_written += written
                            s.filled_records.append(key)
                            log(f"    FILLED   {key}  ({written} cells)")
                        else:
                            s.rows_unchanged += 1

                    # ── Append missing row ────────────────────────────────
                    else:
                        if not ALLOW_APPEND_NEW_ROWS:
                            s.skipped_append_disabled += 1; s.warnings += 1
                            log(f"    SKIPPED  {key}  (append disabled)"); continue

                        new_row = build_new_row(zone_name, sheet_name, month_name, src_df, master_columns,
                                                target_start_year, target_fy_label, summary=s)
                        new_er  = ws.max_row + 1
                        for ci, cn in enumerate(master_columns, 1):
                            ws.cell(row=new_er, column=ci).value = new_row.get(cn)

                        master_index[key] = new_er
                        added_cells = sum(1 for v in new_row.values() if v not in (None, ""))
                        s.rows_added += 1; s.cells_written += added_cells
                        s.filled_records.append(key)
                        log(f"    ADDED    {key}  ({added_cells} cells)")

        # ── Save ──────────────────────────────────────────────────────────────
        log("")
        log("=" * 72)
        total = s.rows_updated + s.rows_added

        if total == 0:
            log("RawData is already up to date. Nothing to write.")
        elif test_mode:
            log(f"TEST MODE – {total} record(s) would be updated. No file saved.")
        else:
            save_path = OUTPUT_FILE or RAWDATA_MASTER_FILE
            log(f"Saving → {save_path}")
            wb.save(save_path)
            log("Saved successfully.")

        # ── Summary ───────────────────────────────────────────────────────────
        log("")
        log("SUMMARY")
        log("-" * 72)
        log(f"Source files read             : {s.source_files_read}")
        log(f"Source sheets processed       : {s.source_sheets_processed}")
        log(f"Source month-records checked  : {s.source_records_seen}")
        log(f"Already complete (skipped)    : {s.skipped_already_complete}")
        log(f"Rows filled (were empty)      : {s.rows_updated}")
        log(f"Rows appended (were missing)  : {s.rows_added}")
        log(f"Rows unchanged                : {s.rows_unchanged}")
        log(f"Cells written                 : {s.cells_written}")
        log(f"Skipped – source has no data  : {s.skipped_empty_source}")
        log(f"Skipped – layout errors       : {s.skipped_layout_error}")
        log(f"Warnings                      : {s.warnings}")
        log(f"Errors                        : {s.errors}")
        log("-" * 72)

        if s.filled_records:
            log("Records filled this run:")
            for r in s.filled_records:
                log(f"  {r}")
        else:
            log("No records needed updating.")

        log("\nDone.")
        return (0 if s.errors == 0 else 1, s, list(_LOG_LINES))

    except FileNotFoundError as e: log(f"FATAL – File not found: {e}"); return (2, s, list(_LOG_LINES))
    except KeyError as e:          log(f"FATAL – Missing column: {e}"); return (3, s, list(_LOG_LINES))
    except ValueError as e:        log(f"FATAL – Data/sheet issue: {e}"); return (4, s, list(_LOG_LINES))
    except Exception as e:
        log("FATAL – Unexpected failure"); log(str(e)); log(traceback.format_exc()); return (99, s, list(_LOG_LINES))
    finally:
        _CAPTURE = False


def main(test_mode=TEST_MODE, force_overwrite=False, fiscal_year=None):
    """CLI wrapper – returns just the integer exit code."""
    code, _summary, _log = run_update(
        test_mode=test_mode, force_overwrite=force_overwrite, fiscal_year=fiscal_year
    )
    return code


if __name__ == "__main__":
    p = argparse.ArgumentParser(
        description="Smart-diff update: fills only missing/zero records in RawData.xlsx.")
    p.add_argument("--test",  action="store_true", help="Dry-run – show changes without saving.")
    p.add_argument("--force", action="store_true", help="Overwrite existing values too (full refresh).")
    p.add_argument("--fiscal-year", default=None,
                   help="Fiscal year the zone workbooks belong to, e.g. FY2026/27 "
                        "(default: auto-detect from the master).")
    a = p.parse_args()
    sys.exit(main(test_mode=a.test, force_overwrite=a.force, fiscal_year=a.fiscal_year))
