"""Compile per-zone workbooks in a FY folder into a DataEntry-format RawData file.
Reuses the row/column mapping from update_rawdata_master.py, fills Year/Fiscal Year
(which that script omits), and writes a clean single-sheet workbook the importer reads.

Usage: python tools/compile_rawdata.py
(edit the calls at the bottom for other folders/years)
"""
import pandas as pd
from openpyxl import load_workbook, Workbook

import os as _os
# Master template (defines the DataEntry header layout). uploads/ is
# environment-managed and may be cleared, so fall back to a stable copy.
MASTER = next((p for p in (
    r"D:\WebApps\opsapp\uploads\RawData.xlsx",
    r"D:\WebApps\opsapp\dataupdater\RawData.xlsx",
) if _os.path.exists(p)), r"D:\WebApps\opsapp\dataupdater\RawData.xlsx")
ZONES = ["Liwonde", "Mangochi", "Mulanje", "Ngabu", "Zomba"]
NON_DATA = {"Zone Monthly","1st QTR","2nd QTR ","Midyear Summary","3rd QTR","4th QTR",
            "Zone Annual Summary","Targets","Budget","ZA TH JA CALCULATION",
            "COST OF POWER","WATER TARIFF"}
MONTH_COL = {"April":3,"May":4,"June":5,"July":6,"August":7,"September":8,"October":9,
             "November":10,"December":11,"January":12,"February":13,"March":14}
MONTH_NO = {"April":4,"May":5,"June":6,"July":7,"August":8,"September":9,"October":10,
            "November":11,"December":12,"January":1,"February":2,"March":3}
SRC_MAP = [(8,7),(10,8),(11,9),(12,10),(13,11),(16,13),(17,14),(18,15),(19,16),(23,19),
    (28,21),(30,23),(32,25),(34,27),(36,29),(38,31),(40,33),(43,35),(45,37),(48,39),(49,40),
    (50,41),(51,42),(53,43),(54,44),(55,45),(60,49),(62,51),(65,52),(66,53),(67,54),(68,55),
    (71,58),(72,59),(73,60),(74,61),(77,64),(78,65),(79,66),(80,67),(83,70),(84,71),(85,72),
    (86,73),(97,80),(100,81),(101,82),(102,83),(103,84),(106,86),(107,87),(108,88),(109,89),
    (112,91),(113,92),(114,93),(115,94),(118,97),(120,98),(121,99),(124,101),(125,102),
    (126,103),(127,104),(129,106),(130,107),(131,108),(132,109),(134,111),(135,112),(136,113),
    (137,114),(139,116),(140,117),(141,118),(142,119),(152,126),(153,127),(154,128),(155,129),
    (156,130),(157,131),(158,132),(159,133),(160,134),(161,135),(162,136),(163,137),(164,138),
    (166,139),(167,140),(168,141),(169,142),(170,143),(171,144),(172,145),(173,146),(174,147),
    (176,148),(177,149),(178,150),(179,151),(180,152),(181,153),(183,154),(184,155),(185,156),
    (186,157),(188,158),(189,159),(190,160),(191,161),(193,163),(194,164),(195,165),(196,166),
    (198,167),(199,168),(200,169),(201,170),(202,171),(205,173),(206,174),(207,175),(208,176),
    (211,178),(212,179),(213,180),(214,181),(218,184),(219,185),(220,186),(221,187),(224,189),
    (225,190),(226,191),(227,192),(231,195),(232,196),(233,197),(234,198),(237,200),(238,201),
    (239,202),(240,203),(244,206),(245,207),(252,212),(253,213),(254,214),(255,215),(256,216),
    (260,219),(261,220)]

# The original map only carried component rows; the DataEntry master computed
# the TOTAL columns by formula (so a fresh workbook left them blank → 0 on
# import). Add the source TOTAL rows explicitly so billing/customer headline
# columns are populated. (source_row, master_col_index)
SRC_MAP += [
    (14, 12),    # TOTAL Vol Billed Postpaid
    (110, 90),   # TOTAL Active Postpaid
    (116, 95),   # TOTAL Active Prepaid
    (117, 96),   # TOTAL Active Customers
    (209, 177),  # TOTAL Cash Coll PP
    (215, 182),  # TOTAL Cash Coll Prepaid
    (216, 183),  # TOTAL Cash Collected
    (222, 188),  # TOTAL Amt Billed PP
    (228, 193),  # TOTAL Amt Billed Prepaid
    (229, 194),  # TOTAL Amount Billed
    (242, 205),  # TOTAL Sales MWK
]

def quarter(mn): return "Q1" if mn in(4,5,6) else "Q2" if mn in(7,8,9) else "Q3" if mn in(10,11,12) else "Q4"
def num(v):
    try:
        if pd.isna(v): return 0
    except Exception: pass
    if isinstance(v,str):
        v=v.strip().replace(",","")
        if v=="": return 0
    try: return float(v)
    except Exception: return 0

def compile_fy(folder, end_year, fy_label, out_path):
    mwb = load_workbook(MASTER, read_only=True)
    ws = mwb["DataEntry"]
    row1 = next(ws.iter_rows(min_row=1,max_row=1,values_only=True))
    row2 = next(ws.iter_rows(min_row=2,max_row=2,values_only=True))
    headers = list(row2); ncol = len(headers); mwb.close()

    out = Workbook(); o = out.active; o.title = "DataEntry"
    for j,v in enumerate(row1, start=1): o.cell(row=1,column=j,value=v)
    for j,v in enumerate(headers, start=1): o.cell(row=2,column=j,value=v)

    rrow = 3; nrows = 0; schemes = set()
    for zone in ZONES:
        zpath = f"{folder}\\{zone}.xlsx"
        zx = pd.ExcelFile(zpath)
        for sheet in zx.sheet_names:
            if sheet in NON_DATA: continue
            sdf = pd.read_excel(zpath, sheet_name=sheet, header=None)
            for mname, scol in MONTH_COL.items():
                if scol >= sdf.shape[1]: continue
                probe = sdf.iloc[8, scol] if sdf.shape[0] > 8 else None
                blank = probe is None or (isinstance(probe,str) and probe.strip()=="") or (not isinstance(probe,str) and num(probe)==0)
                if blank: continue
                mn = MONTH_NO[mname]
                o.cell(row=rrow,column=1,value=zone)
                o.cell(row=rrow,column=2,value=sheet)
                o.cell(row=rrow,column=3,value=fy_label)
                o.cell(row=rrow,column=4,value=(end_year-1 if mn>=4 else end_year))
                o.cell(row=rrow,column=5,value=mn)
                o.cell(row=rrow,column=6,value=mname)
                o.cell(row=rrow,column=7,value=quarter(mn))
                for srow, rcol in SRC_MAP:
                    if rcol >= ncol: continue
                    try: val = num(sdf.iloc[srow, scol])
                    except Exception: val = 0
                    o.cell(row=rrow, column=rcol+1, value=val)
                rrow += 1; nrows += 1; schemes.add((zone,sheet))
    out.save(out_path)
    print(f"{fy_label}: wrote {nrows} rows, {len(schemes)} schemes -> {out_path}")
    return out_path

if __name__ == "__main__":
    compile_fy(r"D:\WebApps\opsapp\FY2324", 2024, "FY2023/24", r"D:\WebApps\opsapp\FY2324\RawData_FY2023-24.xlsx")
    compile_fy(r"D:\WebApps\opsapp\FY2425", 2025, "FY2024/25", r"D:\WebApps\opsapp\FY2425\RawData_FY2024-25.xlsx")
